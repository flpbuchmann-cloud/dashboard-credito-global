"""
Extrator de Data Pack XLSX do HSBC Holdings plc.

HSBC publica trimestralmente um "data-pack-<PERIOD>.xlsx" contendo todas as
planilhas-chave do grupo. Cada arquivo traz os ultimos 5 trimestres em
colunas independentes (nao cumulativos), o que dispensa qualquer
desacumulacao.

HSBC reporta em USD desde 2018 - nao ha conversao de moeda.

Estrategia:
- Le todos os data-pack-*.xlsx na pasta Documentos
- Para cada arquivo extrai dados das abas "Group income statement",
  "Group balance sheet", "Group NIM" e "Credit risk"
- Cada arquivo fornece ate 5 trimestres standalone. Arquivos mais recentes
  sobrescrevem dados de arquivos antigos (numeros reapresentados).
- Salva supplement_data.json na pasta destino.
"""

from __future__ import annotations

import os
import re
import json
import glob
from datetime import datetime
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

_MONTHS = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "may": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}


def _date_to_str(val) -> str | None:
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        s = val.strip()
        # '31-Dec-21' or '31 Dec 21' or '31-Dec-2021'
        m = re.match(r"^(\d{1,2})[-\s]([A-Za-z]{3})[-\s](\d{2,4})$", s)
        if m:
            d, mon, y = m.group(1), m.group(2).lower(), m.group(3)
            mm = _MONTHS.get(mon)
            if not mm:
                return None
            if len(y) == 2:
                y = "20" + y
            return f"{y}-{mm}-{int(d):02d}"
        # 'YYYY-MM-DD'
        if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
            return s
    return None


def _date_to_trimestre(date_str: str) -> str:
    y, m, _ = date_str.split("-")
    q = {"03": 1, "06": 2, "09": 3, "12": 4}.get(m, 4)
    return f"{q}Q{int(y) % 100:02d}"


def _find_row(ws, label_regex: str, col: int = 1, max_row: int | None = None) -> int | None:
    """Acha a primeira linha em que a celula [col] casa com o regex (case-insensitive)."""
    rx = re.compile(label_regex, re.IGNORECASE)
    last = max_row or ws.max_row
    for r in range(1, last + 1):
        v = ws.cell(r, col).value
        if isinstance(v, str) and rx.search(v):
            return r
    return None


def _header_date_columns(ws, max_scan_rows: int = 80) -> dict[int, str]:
    """
    Encontra a linha de cabecalho "Quarter ended" + datas e retorna
    {col_idx: 'YYYY-MM-DD'} apenas para colunas com datas reais (ignora YTD).

    A aba traz "Quarter ended" seguido de "Year to date" na mesma linha.
    Pegamos apenas as primeiras datas (ate aparecer coluna None ou ate 5 datas).
    """
    # Localiza a linha que contem "Quarter ended" ou "Quarter to date"
    header_row = None
    for r in range(1, max_scan_rows + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and ("quarter ended" in v.lower() or "quarter to date" in v.lower()):
                header_row = r + 1  # datas na linha seguinte
                break
        if header_row:
            break
    if not header_row:
        return {}
    cols: dict[int, str] = {}
    # Le ate encontrar uma celula None (separador antes do YTD) ou 5 datas
    for c in range(2, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        ds = _date_to_str(v)
        if ds:
            cols[c] = ds
            if len(cols) >= 5:
                break
        elif v is None and cols:
            break
    return cols


def _balance_sheet_date_columns(ws) -> dict[int, str]:
    """
    Para a aba Group balance sheet: le datas na linha 6 ou 7 (varia
    conforme versao do arquivo).
    """
    for r in range(4, 10):
        row = [ws.cell(r, c).value for c in range(2, 8)]
        dates = [(i + 2, _date_to_str(v)) for i, v in enumerate(row) if _date_to_str(v)]
        if len(dates) >= 2:
            return {c: d for c, d in dates}
    return {}


def _num(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        t = v.strip().replace(",", "")
        if not t or t in ("-", "n/a", "N/A", "na"):
            return None
        try:
            return float(t)
        except ValueError:
            return None
    return None


def _row_values(ws, row: int, cols: dict[int, str]) -> dict[str, float]:
    out: dict[str, float] = {}
    for c, ds in cols.items():
        v = _num(ws.cell(row, c).value)
        if v is not None:
            out[ds] = v
    return out


# ---------------------------------------------------------------------------
# Extracao por aba
# ---------------------------------------------------------------------------

def _extract_income_statement(wb) -> dict[str, dict]:
    """
    Retorna {date_str: {nii, total_income, total_opex, credit_impairment,
                         profit_before_tax, net_income}}.
    Apenas colunas de "Quarter ended" (standalone).
    """
    if "Group income statement" not in wb.sheetnames:
        return {}
    ws = wb["Group income statement"]
    cols = _header_date_columns(ws)
    if not cols:
        return {}

    row_nii = _find_row(ws, r"^Net interest income$")
    row_rev = _find_row(ws, r"^Net operating income before change")
    row_ecl = _find_row(ws, r"^Change in expected credit losses")
    row_opex = _find_row(ws, r"^Total operating expenses$")
    row_pbt = _find_row(ws, r"Profit.*loss.*before tax|^Profit before tax$")
    row_pat = _find_row(ws, r"Profit.*loss.*after tax|^Profit after tax$")

    data_by_date: dict[str, dict] = {ds: {} for ds in cols.values()}

    def _fill(row, key):
        if row is None:
            return
        for c, ds in cols.items():
            v = _num(ws.cell(row, c).value)
            if v is not None:
                data_by_date[ds][key] = v

    _fill(row_nii, "nii")
    _fill(row_rev, "total_income")
    _fill(row_ecl, "credit_impairment")
    _fill(row_opex, "total_opex")
    _fill(row_pbt, "profit_before_tax")
    _fill(row_pat, "net_income")

    return data_by_date


def _extract_balance_sheet(wb) -> dict[str, dict]:
    """
    Retorna {date_str: {total_assets, gross_loans, customer_deposits,
                         rwa, cet1_capital, cet1_ratio, total_capital_ratio,
                         leverage_ratio}}.
    """
    if "Group balance sheet" not in wb.sheetnames:
        return {}
    ws = wb["Group balance sheet"]
    cols = _balance_sheet_date_columns(ws)
    if not cols:
        return {}

    row_total_assets = _find_row(ws, r"^Total assets$")
    row_customer_acc = _find_row(ws, r"^Customer accounts$")
    row_gross_loans = _find_row(ws, r"Loans and advances to customers \(gross\)")
    row_rwa = _find_row(ws, r"^Risk.weighted assets$")
    row_cet1_cap = _find_row(ws, r"^Common equity tier 1 capital$")
    row_cet1_ratio = _find_row(ws, r"^Common equity tier 1 ratio$")
    row_total_cap_ratio = _find_row(ws, r"^Total capital ratio$")
    row_leverage = _find_row(ws, r"^Leverage [Rr]atio$")

    out: dict[str, dict] = {ds: {} for ds in cols.values()}

    def _fill(row, key):
        if row is None:
            return
        for c, ds in cols.items():
            v = _num(ws.cell(row, c).value)
            if v is not None:
                out[ds][key] = v

    _fill(row_total_assets, "total_assets")
    _fill(row_customer_acc, "customer_deposits")
    _fill(row_gross_loans, "gross_loans")
    _fill(row_rwa, "rwa")
    _fill(row_cet1_cap, "cet1_capital")
    _fill(row_cet1_ratio, "cet1_ratio")
    _fill(row_total_cap_ratio, "total_capital_ratio")
    _fill(row_leverage, "leverage_ratio")
    return out


def _extract_nim(wb) -> dict[str, dict]:
    """
    Retorna {date_str: {avg_earning_assets, avg_loans, avg_customer_deposits, nim}}.
    Apenas colunas "Quarter to date".
    """
    if "Group NIM" not in wb.sheetnames:
        return {}
    ws = wb["Group NIM"]
    cols = _header_date_columns(ws)
    if not cols:
        return {}

    row_avg_loans = _find_row(ws, r"^Loans and advances to customers$")
    row_avg_earning = _find_row(ws, r"^Total interest.earning assets$")
    row_customer_acc = _find_row(ws, r"^Customer accounts$")
    row_nim = _find_row(ws, r"^Net interest margin \(%\)")

    out: dict[str, dict] = {ds: {} for ds in cols.values()}

    def _fill(row, key):
        if row is None:
            return
        for c, ds in cols.items():
            v = _num(ws.cell(row, c).value)
            if v is not None:
                out[ds][key] = v

    _fill(row_avg_earning, "avg_earning_assets")
    _fill(row_avg_loans, "avg_loans")
    _fill(row_customer_acc, "avg_customer_deposits")
    _fill(row_nim, "nim")

    return out


def _extract_credit_risk(wb, period_date: str | None) -> dict:
    """
    A aba Credit risk traz uma foto do credito na data do arquivo (nao 5
    trimestres). Le row 10 ("Loans and advances to customers at amortised cost")
    para obter Stage 1/2/3 gross e ECL allowance.

    Retorna {carteira_credito_bruta, provisao_acumulada, npl, coverage_ratio}.
    """
    if "Credit risk" not in wb.sheetnames:
        return {}
    ws = wb["Credit risk"]
    row = _find_row(ws, r"^Loans and advances to customers at amortised cost", max_row=30)
    if row is None:
        return {}

    # Layout esperado (colunas B..L):
    # B=Stage1 gross, C=Stage2 gross, D=Stage3 gross, E=POCI, F=Total gross
    # G=Stage1 ECL, H=Stage2 ECL, I=Stage3 ECL, J=POCI ECL, K=Total ECL
    total_gross = _num(ws.cell(row, 6).value)
    stage3_gross = _num(ws.cell(row, 4).value)
    total_ecl = _num(ws.cell(row, 11).value)

    out: dict = {}
    if total_gross is not None:
        out["carteira_credito_bruta"] = total_gross
    if total_ecl is not None:
        # Allowance vem negativo no data pack - tornamos positivo
        out["provisao_acumulada"] = abs(total_ecl)
    if stage3_gross is not None:
        out["npl"] = stage3_gross
    if stage3_gross and total_ecl:
        stage3_ecl = _num(ws.cell(row, 9).value)
        if stage3_ecl is not None and stage3_gross > 0:
            out["coverage_ratio"] = round(abs(stage3_ecl) / stage3_gross, 6)
    return out


# ---------------------------------------------------------------------------
# Orquestracao
# ---------------------------------------------------------------------------

def _file_period(filename: str) -> str | None:
    m = re.search(r"data-pack-(FY|H1|Q1|Q3)-(\d{4})\.xlsx", filename)
    if not m:
        return None
    tipo, yr = m.group(1), m.group(2)
    end = {"Q1": "03-31", "H1": "06-30", "Q3": "09-30", "FY": "12-31"}[tipo]
    return f"{yr}-{end}"


def _build_record(date_str: str, is_data: dict, bs_data: dict,
                  nim_data: dict, cr_data: dict, fonte: str) -> dict:
    return {
        "periodo": date_str,
        "trimestre": _date_to_trimestre(date_str),
        "moeda_original": "USD",
        "fx_rate": 1.0,
        "fonte": fonte,
        "avg_balances": {
            "avg_total_assets": bs_data.get("total_assets"),
            "avg_earning_assets": nim_data.get("avg_earning_assets"),
            "avg_loans": nim_data.get("avg_loans"),
            "avg_total_deposits": nim_data.get("avg_customer_deposits") or bs_data.get("customer_deposits"),
        },
        "yields_rates": {
            "nim": nim_data.get("nim"),
            "asset_yield": None,
            "interest_spread": None,
        },
        "capital": {
            "cet1_ratio": bs_data.get("cet1_ratio"),
            "cet1_capital": bs_data.get("cet1_capital"),
            "rwa_standardized": bs_data.get("rwa"),
            "leverage_ratio": bs_data.get("leverage_ratio"),
            "lcr": None,
            "nsfr": None,
            "total_capital_ratio": bs_data.get("total_capital_ratio"),
        },
        "credit_quality": {
            "carteira_credito_bruta": cr_data.get("carteira_credito_bruta") or bs_data.get("gross_loans"),
            "provisao_acumulada": cr_data.get("provisao_acumulada"),
            "npl": cr_data.get("npl"),
            "coverage_ratio": cr_data.get("coverage_ratio"),
            "nco_total": None,
        },
        "income_statement": {
            "total_income": is_data.get("total_income"),
            "nii": is_data.get("nii"),
            "total_opex": is_data.get("total_opex"),
            "credit_impairment": is_data.get("credit_impairment"),
            "profit_before_tax": is_data.get("profit_before_tax"),
            "net_income": is_data.get("net_income"),
        },
    }


def extrair_supplement_hsbc(pasta_docs: str, pasta_destino: str) -> list[dict]:
    """Extrai data pack XLSX do HSBC e salva supplement_data.json."""
    xlsx_files = sorted(
        glob.glob(os.path.join(pasta_docs, "data-pack-*.xlsx")),
        key=lambda p: (_file_period(os.path.basename(p)) or "")
    )
    if not xlsx_files:
        print(f"[WARN] Nenhum data pack encontrado em {pasta_docs}")
        return []

    print(f"[INFO] Encontrados {len(xlsx_files)} data packs HSBC")

    # Agrega por trimestre. Arquivos mais recentes sobrescrevem antigos.
    by_date: dict[str, dict] = {}

    for path in xlsx_files:
        fname = os.path.basename(path)
        period = _file_period(fname)
        print(f"  Processando {fname} (periodo {period})...")
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        except Exception as e:
            print(f"    [ERRO] {e}")
            continue

        try:
            is_map = _extract_income_statement(wb)
            bs_map = _extract_balance_sheet(wb)
            nim_map = _extract_nim(wb)
            cr_snapshot = _extract_credit_risk(wb, period)
        finally:
            wb.close()

        # Consolida todas as datas vistas neste arquivo
        all_dates = set(is_map) | set(bs_map) | set(nim_map)
        for ds in all_dates:
            merged = by_date.get(ds, {
                "is": {}, "bs": {}, "nim": {}, "cr": {}, "fonte": fname
            })
            # Sempre sobrescreve (arquivo mais recente ganha pois iteramos ordenado)
            if is_map.get(ds):
                merged["is"] = is_map[ds]
            if bs_map.get(ds):
                merged["bs"] = bs_map[ds]
            if nim_map.get(ds):
                merged["nim"] = nim_map[ds]
            merged["fonte"] = fname
            by_date[ds] = merged

        # Credit risk snapshot aplica-se ao periodo do arquivo apenas
        if cr_snapshot and period and period in by_date:
            by_date[period]["cr"] = cr_snapshot
            by_date[period]["fonte"] = fname

    records = [
        _build_record(ds, v["is"], v["bs"], v["nim"], v["cr"], v["fonte"])
        for ds, v in sorted(by_date.items())
    ]

    os.makedirs(pasta_destino, exist_ok=True)
    out_path = os.path.join(pasta_destino, "supplement_data.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)

    print(f"\n[INFO] Salvos {len(records)} trimestres em {out_path}")

    # Tambem gera contas_chave.json no formato do dashboard
    _gerar_contas_chave(records, pasta_destino)

    # Resumo
    print("\n--- Resumo ---")
    for r in records:
        ni = r["income_statement"].get("net_income")
        cet1 = r["capital"].get("cet1_ratio")
        nim = r["yields_rates"].get("nim")
        ta = r["avg_balances"].get("avg_total_assets")
        ni_str = f"${ni:,.0f}m" if ni is not None else "N/A"
        cet1_str = f"{cet1:.1%}" if cet1 is not None else "N/A"
        nim_str = f"{nim:.2%}" if nim is not None else "N/A"
        ta_str = f"${ta/1000:,.0f}bn" if ta is not None else "N/A"
        print(f"  {r['trimestre']}: NI={ni_str}  CET1={cet1_str}  NIM={nim_str}  TA={ta_str}")

    return records


# ---------------------------------------------------------------------------
# contas_chave.json (shim para o dashboard ler HSBC)
# ---------------------------------------------------------------------------

def _gerar_contas_chave(records: list[dict], pasta_destino: str) -> None:
    """
    Converte supplement_data (HSBC, em USD milhoes) para o esquema
    contas_chave.json usado pelo dashboard. Valores sao expandidos para USD.
    """
    M = 1_000_000  # USD millions -> USD
    out = []
    for r in records:
        periodo = r["periodo"]
        ano = int(periodo[:4])
        mes = int(periodo[5:7])
        tipo_prefix = "DFP" if mes == 12 else "ITR"

        is_d = r.get("income_statement", {})
        bs_avg = r.get("avg_balances", {})
        cr = r.get("credit_quality", {})
        cap = r.get("capital", {})
        yr = r.get("yields_rates", {})

        def mul(v):
            return v * M if isinstance(v, (int, float)) else 0.0

        # DRE
        nii = is_d.get("nii") or 0
        total_income = is_d.get("total_income") or 0
        non_int = (total_income - nii) if (total_income and nii) else 0
        opex = is_d.get("total_opex") or 0
        ecl = is_d.get("credit_impairment") or 0
        pbt = is_d.get("profit_before_tax") or 0
        ni = is_d.get("net_income") or 0
        ir = pbt - ni
        dre_contas = {
            "receita_liquida": mul(total_income),
            "nii": mul(nii),
            "receita_juros": mul(nii),
            "receita_nao_juros": mul(non_int),
            "despesa_juros": 0.0,
            "despesas_operacionais": mul(abs(opex)),
            "provisao_credito": mul(abs(ecl)),
            "compensacao": 0.0,
            "marketing": 0.0,
            "depreciacao_amortizacao": 0.0,
            "ebit": mul(pbt),
            "lucro_antes_ir": mul(pbt),
            "ir_csll": mul(ir),
            "lucro_liquido": mul(ni),
            "resultado_equivalencia": 0.0,
        }
        out.append({"periodo": periodo, "tipo": f"{tipo_prefix}_dre", "ano": ano, "contas": dre_contas})

        # BPA
        total_assets = bs_avg.get("avg_total_assets") or 0
        earn = bs_avg.get("avg_earning_assets") or 0
        loans = cr.get("carteira_credito_bruta") or bs_avg.get("avg_loans") or 0
        prov = cr.get("provisao_acumulada") or 0
        npl = cr.get("npl") or 0
        bpa_contas = {
            "ativo_total": mul(total_assets),
            "ativo_circulante": 0.0,
            "ativo_nao_circulante": 0.0,
            "caixa": 0.0,
            "contas_a_receber": 0.0,
            "depositos_em_bancos": 0.0,
            "earning_assets": mul(earn),
            "emprestimos_concedidos": mul(loans),
            "carteira_credito_bruta": mul(loans),
            "provisao_acumulada": mul(prov),
            "npl": mul(npl),
            "investimentos_titulos": 0.0,
            "imobilizado": 0.0,
            "intangivel": 0.0,
            "lucros_retidos": 0.0,
        }
        out.append({"periodo": periodo, "tipo": f"{tipo_prefix}_bpa", "ano": ano, "contas": bpa_contas})

        # BPP
        deposits = bs_avg.get("avg_total_deposits") or 0
        cet1_cap = cap.get("cet1_capital") or 0
        bpp_contas = {
            "passivo_total": mul(total_assets - (cet1_cap if cet1_cap else 0)),
            "passivo_circulante": 0.0,
            "passivo_nao_circulante": 0.0,
            "depositos": mul(deposits),
            "depositos_interest_bearing_domestic": 0.0,
            "depositos_interest_bearing_foreign": 0.0,
            "depositos_noninterest_bearing": 0.0,
            "short_term_borrowings": 0.0,
            "emprestimos_cp": 0.0,
            "emprestimos_lp": 0.0,
            "outras_obrigacoes_cp": 0.0,
            "outras_obrigacoes_lp": 0.0,
            "provisoes_cp": 0.0,
            "provisoes_lp": 0.0,
            "patrimonio_liquido": mul(cet1_cap),
            "capital_social": 0.0,
        }
        out.append({"periodo": periodo, "tipo": f"{tipo_prefix}_bpp", "ano": ano, "contas": bpp_contas})

        # DFC (vazio - dashboard tolerante)
        dfc_contas = {k: 0.0 for k in [
            "caixa_gerado_operacoes", "capex", "fco", "fci", "fcf",
            "dividendos_pagos", "recompra_acoes", "juros_pagos",
            "juros_emprestimos_dfc", "captacao_divida", "amortizacao_divida",
            "depreciacao_amortizacao", "var_ativos_passivos"
        ]}
        out.append({"periodo": periodo, "tipo": f"{tipo_prefix}_dfc", "ano": ano, "contas": dfc_contas})

        # PSD
        psd_contas = {
            "acoes_diluidas": 0.0, "acoes_outstanding": 0.0,
            "dividendo_por_acao": 0.0, "lpa_diluido": 0.0,
        }
        out.append({"periodo": periodo, "tipo": f"{tipo_prefix}_psd", "ano": ano, "contas": psd_contas})

    p = os.path.join(pasta_destino, "contas_chave.json")
    with open(p, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f"[INFO] contas_chave.json salvo ({len(out)} registros)")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    pasta_docs = r"G:\Meu Drive\Análise de Crédito Financeiras\HSBC\Documentos"
    pasta_destino = r"G:\Meu Drive\Análise de Crédito Financeiras\HSBC\Dados_EDGAR"

    recs = extrair_supplement_hsbc(pasta_docs, pasta_destino)
    print("\n--- Ultimas 2 entradas ---")
    for r in recs[-2:]:
        print(json.dumps(r, indent=2, ensure_ascii=False))

    # ratings.json / ri_website.json
    pasta_hsbc = r"G:\Meu Drive\Análise de Crédito Financeiras\HSBC"
    with open(os.path.join(pasta_hsbc, "ratings.json"), "w", encoding="utf-8") as f:
        json.dump({
            "moodys": "A1",
            "sp": "A+",
            "fitch": "AA-",
            "ticker": "HSBC",
            "data_consulta": datetime.now().isoformat(timespec="seconds"),
            "fonte": "Manual (agencias de rating publicas)"
        }, f, indent=2, ensure_ascii=False)
    with open(os.path.join(pasta_hsbc, "ri_website.json"), "w", encoding="utf-8") as f:
        json.dump({
            "ri_url": "https://www.hsbc.com/investors/results-and-announcements",
            "ticker": "HSBC",
            "data_consulta": datetime.now().isoformat(timespec="seconds"),
            "fonte": "manual"
        }, f, indent=2, ensure_ascii=False)
    print("\n[INFO] ratings.json e ri_website.json gerados em", pasta_hsbc)
