"""
Enrich Barclays supplement_data.json and contas_chave.json with data
extracted from the official Financial Tables XLSX files.

Extracts:
  - Group liquidity pool (HQLA) per quarter
  - LCR HQLA (£bn) and NSFR details (annual)
  - Deposit funding breakdown (annual)
  - NIM by division (annual + quarterly)
  - Group Qrtly P&L, metrics, balance sheet, funding

All £bn values are converted to USD millions using the fx_rate in each
supplement entry.
"""

import json
import re
import openpyxl
from pathlib import Path

# ── Paths ──────────────────────────────────────────────────────────────
BCS_DIR = Path("G:/Meu Drive/Análise de Crédito Financeiras/BCS")
DOCS_DIR = BCS_DIR / "Documentos"
DADOS_DIR = BCS_DIR / "Dados_EDGAR"
SUPPLEMENT_PATH = DADOS_DIR / "supplement_data.json"
CONTAS_PATH = DADOS_DIR / "contas_chave.json"

# Quarter label → period end date
QTR_TO_DATE = {}
for y in range(2020, 2030):
    yy = str(y)[2:]
    QTR_TO_DATE[f"Q1{yy}"] = f"{y}-03-31"
    QTR_TO_DATE[f"Q2{yy}"] = f"{y}-06-30"
    QTR_TO_DATE[f"Q3{yy}"] = f"{y}-09-30"
    QTR_TO_DATE[f"Q4{yy}"] = f"{y}-12-31"


def find_xlsx_files():
    """Return sorted list of financial-tables XLSX files."""
    return sorted(DOCS_DIR.glob("financial-tables-FY-*.xlsx"))


def read_group_qrtly(wb):
    """
    Read the Group Qrtly sheet and return a dict keyed by quarter label
    (e.g. 'Q425') with all extracted row data.
    """
    ws = None
    for name in wb.sheetnames:
        if "qrtly" in name.lower() and "group" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        return {}

    # Find the header row with quarter labels (Q4XX, Q3XX, etc.)
    header_row = None
    qtr_cols = {}  # col_index -> quarter label
    for r in range(1, 6):
        row = ws[r]
        for c in row:
            if c.value and isinstance(c.value, str) and re.match(r"Q[1-4]\d{2}", c.value):
                qtr_cols[c.column] = c.value
        if qtr_cols:
            header_row = r
            break

    if not qtr_cols:
        return {}

    # Row label mapping: row label text -> our key name
    ROW_MAP = {
        "net interest income": "nii",
        "net fee, commission and other income": "non_interest_income",
        "total income": "total_income",
        "operating costs": "operating_costs",
        "uk regulatory levies": "uk_regulatory_levies",
        "litigation and conduct": "litigation_and_conduct",
        "total operating expenses": "total_opex",
        "other net (expenses)/income": "other_net_income",
        "profit before impairment": "profit_before_impairment",
        "credit impairment charges": "credit_impairment",
        "profit before tax": "profit_before_tax",
        "tax charges": "tax_charges",
        "profit after tax": "profit_after_tax",
        "non-controlling interests": "nci",
        "other equity instrument holders": "other_equity_holders",
        "attributable profit": "attributable_profit",
        "return on average tangible shareholders' equity": "rote",
        "average tangible shareholders' equity (£bn)": "avg_tse_bn",
        "cost: income ratio": "cost_income_ratio",
        "loan loss rate (bps)": "loan_loss_rate_bps",
        "basic earnings per ordinary share": "eps",
        "basic weighted average number of shares (m)": "shares_wavg",
        "period end number of shares (m)": "shares_period_end",
        "period end tangible shareholders' equity (£bn)": "tse_bn",
        "loans and advances to customers at amortised cost": "customer_loans_bn",
        "loans and advances to banks at amortised cost": "bank_loans_bn",
        "debt securities at amortised cost": "debt_securities_bn",
        "loans and advances at amortised cost": "total_loans_advances_bn",
        "loans and advances at amortised cost impairment coverage ratio": "impairment_coverage",
        "total assets": "total_assets_bn",
        "deposits at amortised cost": "deposits_bn",
        "tangible net asset value per share": "tnav_per_share",
        "common equity tier 1 ratio": "cet1_ratio",
        "common equity tier 1 capital": "cet1_capital_bn",
        "risk weighted assets": "rwa_bn",
        "uk leverage ratio": "leverage_ratio",
        "uk leverage exposure": "leverage_exposure_bn",
        "group liquidity pool (£bn)": "hqla_pool_bn",
        "liquidity coverage ratio": "lcr",
        "net stable funding ratio": "nsfr",
        "loan: deposit ratio": "loan_deposit_ratio",
    }

    result = {q: {} for q in qtr_cols.values()}

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
        label_cell = row[1]  # Column B (index 1 in 0-based)
        if not label_cell.value or not isinstance(label_cell.value, str):
            continue

        label_clean = label_cell.value.strip().rstrip(" ")
        # Normalize unicode
        label_lower = label_clean.lower().replace("\u00a3", "£")

        key = None
        for pattern, k in ROW_MAP.items():
            if label_lower.startswith(pattern):
                key = k
                break
        if key is None:
            continue

        for col_idx, qtr_label in qtr_cols.items():
            cell = row[col_idx - 1]  # 0-based index
            if cell.value is not None:
                result[qtr_label][key] = cell.value

    return result


def read_lcr_nsfr(wb):
    """Read LCR and NSFR sheets (annual point-in-time data)."""
    data = {}

    # LCR
    ws = wb["Liquidity coverage ratio"] if "Liquidity coverage ratio" in wb.sheetnames else None
    if ws:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            label = row[1].value if len(row) > 1 and row[1].value else ""
            if isinstance(label, str) and "hqla" in label.lower():
                if len(row) > 2 and row[2].value is not None:
                    data["lcr_hqla_bn_current"] = row[2].value
                if len(row) > 3 and row[3].value is not None:
                    data["lcr_hqla_bn_prior"] = row[3].value

    # NSFR
    ws = wb["Net stable funding ratio"] if "Net stable funding ratio" in wb.sheetnames else None
    if ws:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            label = row[1].value if len(row) > 1 and row[1].value else ""
            if isinstance(label, str):
                lbl = label.lower()
                if "available stable" in lbl:
                    if len(row) > 2 and row[2].value:
                        data["nsfr_asf_bn_current"] = row[2].value
                    if len(row) > 3 and row[3].value:
                        data["nsfr_asf_bn_prior"] = row[3].value
                elif "required stable" in lbl:
                    if len(row) > 2 and row[2].value:
                        data["nsfr_rsf_bn_current"] = row[2].value
                    if len(row) > 3 and row[3].value:
                        data["nsfr_rsf_bn_prior"] = row[3].value

    return data


def read_deposit_funding(wb):
    """Read Deposit funding sheet (annual)."""
    ws = wb["Deposit funding"] if "Deposit funding" in wb.sheetnames else None
    if not ws:
        return {}

    data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        label = row[1].value if len(row) > 1 and row[1].value else ""
        if not isinstance(label, str):
            continue
        lbl = label.strip().lower()
        if "barclays group" in lbl or "total" in lbl:
            # Row 12: Group totals
            vals = {c.column: c.value for c in row if c.value is not None}
            # Col 3 = loans, Col 4 = deposits, Col 5 = LDR current
            if 3 in vals:
                data["group_loans_bn"] = vals[3]
            if 4 in vals:
                data["group_deposits_bn"] = vals[4]
            if 5 in vals:
                data["group_ldr_pct"] = vals[5]

    return data


def read_margins(wb):
    """Read Margins and balances sheet for quarterly NIM."""
    ws = wb["Margins and balances"] if "Margins and balances" in wb.sheetnames else None
    if not ws:
        return {}

    # Find quarterly section (row 16+)
    data = {}
    in_nim_section = False
    qtr_cols = {}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        label = row[1].value if len(row) > 1 and row[1].value else ""
        if isinstance(label, str) and "quarterly" in label.lower():
            # Next row with Q labels
            continue

        # Check for quarter headers
        for c in row:
            if c.value and isinstance(c.value, str) and re.match(r"Q[1-4]\d{2}", c.value):
                qtr_cols[c.column] = c.value

        if isinstance(label, str) and "net interest margin" in label.lower() and "%" in str(row[2].value if len(row) > 2 else ""):
            in_nim_section = True
            continue

        if in_nim_section and isinstance(label, str):
            lbl = label.strip().lower()
            if "group" in lbl and "excluding" not in lbl:
                # Group NIM - take it
                for col_idx, qtr in qtr_cols.items():
                    cell = row[col_idx - 1]
                    if cell.value is not None:
                        if qtr not in data:
                            data[qtr] = {}
                        data[qtr]["group_nim_pct"] = cell.value
            elif "excluding" in lbl:
                for col_idx, qtr in qtr_cols.items():
                    cell = row[col_idx - 1]
                    if cell.value is not None:
                        if qtr not in data:
                            data[qtr] = {}
                        data[qtr]["nim_ex_ib_ho_pct"] = cell.value

    # Also get the annual Group NIM from rows 6-13
    for row in ws.iter_rows(min_row=6, max_row=13, values_only=False):
        label = row[1].value if len(row) > 1 and row[1].value else ""
        if isinstance(label, str) and "group" in label.lower() and "net interest" in label.lower():
            # Col 3 = current year NII, col 6 = prior year NII
            if len(row) > 2 and row[2].value:
                data["annual_group_nii_m"] = row[2].value

    return data


def read_liquidity_pool(wb):
    """Read Group liquidity pool sheet for HQLA composition."""
    ws = wb["Group liquidity pool"] if "Group liquidity pool" in wb.sheetnames else None
    if not ws:
        return {}

    data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        label = row[1].value if len(row) > 1 and row[1].value else ""
        if not isinstance(label, str):
            continue
        lbl = label.strip().lower()
        if "total as at" in lbl and "2025" in lbl:
            vals = {c.column: c.value for c in row if c.value is not None and isinstance(c.value, (int, float))}
            # Col 3=Cash, 4=L1, 5=L2A, 6=L2B, 7=HQLA total, 9=Liquidity pool
            data["hqla_cash_bn"] = vals.get(3, 0)
            data["hqla_l1_bn"] = vals.get(4, 0)
            data["hqla_l2a_bn"] = vals.get(5, 0)
            data["hqla_l2b_bn"] = vals.get(6, 0)
            data["hqla_total_bn"] = vals.get(7, 0)
            data["liquidity_pool_bn"] = vals.get(9, 0)

    return data


def enrich_supplement(supplement, all_qtr_data, lcr_nsfr, deposit_funding, margins, liq_pool):
    """Update supplement entries with extracted XLSX data."""
    # Build a lookup by periodo
    qtr_lookup = {}
    for qtr_label, qdata in all_qtr_data.items():
        date_str = QTR_TO_DATE.get(qtr_label)
        if date_str:
            qtr_lookup[date_str] = (qtr_label, qdata)

    margin_lookup = {}
    for qtr_label, mdata in margins.items():
        if isinstance(qtr_label, str) and re.match(r"Q[1-4]\d{2}", qtr_label):
            date_str = QTR_TO_DATE.get(qtr_label)
            if date_str:
                margin_lookup[date_str] = mdata

    updated_count = 0

    for entry in supplement:
        periodo = entry["periodo"]
        fx = entry.get("fx_rate", 1.0)

        if periodo in qtr_lookup:
            qtr_label, qdata = qtr_lookup[periodo]
            updated_count += 1

            # ── Capital: HQLA pool ──
            cap = entry.setdefault("capital", {})
            if "hqla_pool_bn" in qdata:
                # Convert £bn to USD millions: £bn * fx_rate * 1000
                cap["hqla_pool"] = round(qdata["hqla_pool_bn"] * fx * 1000, 1)
                cap["hqla_pool_gbp_bn"] = qdata["hqla_pool_bn"]

            # ── Capital: update CET1 etc. from XLSX if available ──
            if "cet1_ratio" in qdata:
                cap["cet1_ratio"] = qdata["cet1_ratio"]
            if "cet1_capital_bn" in qdata:
                cap["cet1_capital"] = round(qdata["cet1_capital_bn"] * fx * 1000, 1)
            if "rwa_bn" in qdata:
                cap["rwa_standardized"] = round(qdata["rwa_bn"] * fx * 1000, 1)
            if "leverage_ratio" in qdata:
                cap["leverage_ratio"] = qdata["leverage_ratio"]
            if "lcr" in qdata:
                # LCR is stored as e.g. 1.7 meaning 170%
                cap["lcr"] = qdata["lcr"]
            if "nsfr" in qdata:
                cap["nsfr"] = qdata["nsfr"]

            # ── Yields/Rates ──
            yr = entry.setdefault("yields_rates", {})
            if "rote" in qdata:
                yr["rote"] = qdata["rote"]
            if "cost_income_ratio" in qdata:
                yr["efficiency_ratio"] = qdata["cost_income_ratio"]
            if "loan_loss_rate_bps" in qdata:
                yr["loan_loss_rate"] = qdata["loan_loss_rate_bps"] / 10000
                entry.setdefault("credit_quality", {})["loan_loss_rate_bps"] = qdata["loan_loss_rate_bps"]

            # ── Balance sheet items ──
            ab = entry.setdefault("avg_balances", {})
            if "deposits_bn" in qdata:
                ab["deposits_at_amortised_cost_gbp_bn"] = qdata["deposits_bn"]
            if "customer_loans_bn" in qdata:
                ab["customer_loans_gbp_bn"] = qdata["customer_loans_bn"]
            if "total_assets_bn" in qdata:
                ab["total_assets_gbp_bn"] = qdata["total_assets_bn"]
            if "loan_deposit_ratio" in qdata:
                ab["loan_to_deposit"] = qdata["loan_deposit_ratio"]

            # ── Funding and liquidity ──
            fl = entry.setdefault("funding_liquidity", {})
            if "hqla_pool_bn" in qdata:
                fl["liquidity_pool_gbp_bn"] = qdata["hqla_pool_bn"]
                fl["liquidity_pool_usd_m"] = round(qdata["hqla_pool_bn"] * fx * 1000, 1)
            if "lcr" in qdata:
                fl["lcr"] = qdata["lcr"]
            if "nsfr" in qdata:
                fl["nsfr"] = qdata["nsfr"]
            if "loan_deposit_ratio" in qdata:
                fl["loan_deposit_ratio"] = qdata["loan_deposit_ratio"]

            # ── Margins ──
            if periodo in margin_lookup:
                mdata = margin_lookup[periodo]
                if "nim_ex_ib_ho_pct" in mdata:
                    yr["nim"] = mdata["nim_ex_ib_ho_pct"] / 100
                    yr["nim_ex_ib_pct"] = mdata["nim_ex_ib_ho_pct"]

    print(f"  Updated {updated_count} supplement entries with XLSX data")
    return supplement


def enrich_contas_chave(contas, all_qtr_data):
    """Add HQLA to contas_chave bpa entries (in USD, full units)."""
    qtr_lookup = {}
    for qtr_label, qdata in all_qtr_data.items():
        date_str = QTR_TO_DATE.get(qtr_label)
        if date_str:
            qtr_lookup[date_str] = qdata

    # We need fx_rates from supplement
    with open(SUPPLEMENT_PATH, "r", encoding="utf-8") as f:
        supplement = json.load(f)
    fx_map = {e["periodo"]: e.get("fx_rate", 1.0) for e in supplement}

    updated = 0
    for entry in contas:
        periodo = entry["periodo"]
        if entry["tipo"] != "ITR_bpa":
            continue
        if periodo not in qtr_lookup:
            continue

        qdata = qtr_lookup[periodo]
        fx = fx_map.get(periodo, 1.0)

        if "hqla_pool_bn" in qdata:
            # Store in full USD units (like other contas_chave values)
            entry["contas"]["hqla_pool"] = round(qdata["hqla_pool_bn"] * fx * 1e9, 0)
            updated += 1

        if "deposits_bn" in qdata:
            entry["contas"]["depositos_amortised_cost"] = round(qdata["deposits_bn"] * fx * 1e9, 0)

        if "total_assets_bn" in qdata:
            entry["contas"]["ativo_total"] = round(qdata["total_assets_bn"] * fx * 1e9, 0)

        if "customer_loans_bn" in qdata:
            entry["contas"]["emprestimos_concedidos"] = round(qdata["customer_loans_bn"] * fx * 1e9, 0)

    print(f"  Updated {updated} contas_chave BPA entries with HQLA")
    return contas


def main():
    print("=" * 60)
    print("Barclays XLSX Enrichment")
    print("=" * 60)

    # ── Step 1: Extract from all FY XLSX files ──
    xlsx_files = find_xlsx_files()
    print(f"\nFound {len(xlsx_files)} FY XLSX files:")
    for f in xlsx_files:
        print(f"  {f.name}")

    all_qtr_data = {}  # qtr_label -> data dict

    for xlsx_path in xlsx_files:
        print(f"\n── Processing {xlsx_path.name} ──")
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

        qrtly = read_group_qrtly(wb)
        print(f"  Group Qrtly: {len(qrtly)} quarters: {list(qrtly.keys())}")

        for qtr_label, qdata in qrtly.items():
            # Only keep if we have a valid date mapping
            if qtr_label in QTR_TO_DATE:
                # Merge (later files overwrite earlier for same quarter)
                if qtr_label not in all_qtr_data:
                    all_qtr_data[qtr_label] = {}
                all_qtr_data[qtr_label].update(qdata)

    # Print summary of what we extracted
    print(f"\n── Total quarters with data: {len(all_qtr_data)} ──")
    for qtr in sorted(all_qtr_data.keys(), key=lambda x: QTR_TO_DATE.get(x, "")):
        d = all_qtr_data[qtr]
        hqla = d.get("hqla_pool_bn", "N/A")
        nii = d.get("nii", "N/A")
        inc = d.get("total_income", "N/A")
        pbt = d.get("profit_before_tax", "N/A")
        cet1 = d.get("cet1_ratio", "N/A")
        print(f"  {qtr} ({QTR_TO_DATE[qtr]}): HQLA={hqla}bn, NII={nii}m, Income={inc}m, PBT={pbt}m, CET1={cet1}")

    # ── Read additional sheets from FY-2025 ──
    print("\n── Additional data from FY-2025 ──")
    wb25 = openpyxl.load_workbook(
        str(DOCS_DIR / "financial-tables-FY-2025.xlsx"), data_only=True
    )
    lcr_nsfr = read_lcr_nsfr(wb25)
    print(f"  LCR/NSFR: {lcr_nsfr}")

    deposit_funding = read_deposit_funding(wb25)
    print(f"  Deposit funding: {deposit_funding}")

    margins = read_margins(wb25)
    print(f"  Margins quarters: {[k for k in margins.keys() if k.startswith('Q')]}")
    for q, m in margins.items():
        if isinstance(q, str) and q.startswith("Q"):
            print(f"    {q}: {m}")

    liq_pool = read_liquidity_pool(wb25)
    print(f"  Liquidity pool composition: {liq_pool}")

    # ── Step 2: Update supplement_data.json ──
    print("\n── Updating supplement_data.json ──")
    with open(SUPPLEMENT_PATH, "r", encoding="utf-8") as f:
        supplement = json.load(f)

    supplement = enrich_supplement(supplement, all_qtr_data, lcr_nsfr, deposit_funding, margins, liq_pool)

    with open(SUPPLEMENT_PATH, "w", encoding="utf-8") as f:
        json.dump(supplement, f, indent=2, ensure_ascii=False)
    print(f"  Saved {SUPPLEMENT_PATH}")

    # ── Step 3: Update contas_chave.json ──
    print("\n── Updating contas_chave.json ──")
    with open(CONTAS_PATH, "r", encoding="utf-8") as f:
        contas = json.load(f)

    contas = enrich_contas_chave(contas, all_qtr_data)

    with open(CONTAS_PATH, "w", encoding="utf-8") as f:
        json.dump(contas, f, indent=2, ensure_ascii=False)
    print(f"  Saved {CONTAS_PATH}")

    # ── Verification ──
    print("\n── Verification: HQLA in supplement ──")
    with open(SUPPLEMENT_PATH, "r", encoding="utf-8") as f:
        supplement = json.load(f)
    for entry in supplement:
        hqla = entry.get("capital", {}).get("hqla_pool", "MISSING")
        hqla_gbp = entry.get("capital", {}).get("hqla_pool_gbp_bn", "N/A")
        fl = entry.get("funding_liquidity", {})
        print(f"  {entry['trimestre']}: HQLA={hqla} USD M (£{hqla_gbp}bn), LCR={fl.get('lcr','N/A')}, NSFR={fl.get('nsfr','N/A')}")

    print("\n── Done! ──")


if __name__ == "__main__":
    main()
