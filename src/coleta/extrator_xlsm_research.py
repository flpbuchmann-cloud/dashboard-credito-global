"""
Extrator generico de modelos XLSM de equity research (multi-broker layouts).

Suporta dois padroes principais:
  A) "MT-style" — datas na linha 1, fy_label "Q4-23" na linha 2, "Hist."/"Forcst." na linha 3
  B) "CI-style" — year header tipo "2023A" na linha 6, quarter "Q1A"/"Q1E" na linha 7

Cada empresa tem um config dict com:
  - sheet_is, sheet_bs, sheet_cf: nomes das abas
  - layout: "A" ou "B"
  - rows_is, rows_bs, rows_cf: dict[contas_chave_key -> linha_xlsm]
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _coerce(v) -> Optional[float]:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s in {"-", "—", "#DIV/0!", "#N/A", "#REF!", "#VALUE!", "n/a", "N/A"}:
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    try:
        f = float(v)
        if f != f:
            return None
        return f
    except (TypeError, ValueError):
        return None


def _normalize_period(year: int, quarter: int) -> str:
    eom = {1: "03-31", 2: "06-30", 3: "09-30", 4: "12-31"}[quarter]
    return f"{year:04d}-{eom}"


def _identify_cols_layout_a(ws, ano_inicio: int) -> list[tuple[int, str]]:
    """Layout MT: row 1 dates, row 2 fy_label 'Q1-08', row 3 Hist/Forcst."""
    out: list[tuple[int, str]] = []
    seen: set[str] = set()
    for c in range(2, ws.max_column + 1):
        d = ws.cell(1, c).value
        fy = str(ws.cell(2, c).value or "")
        hist = str(ws.cell(3, c).value or "")
        if not isinstance(d, datetime):
            continue
        if "Hist" not in hist:
            continue
        if not fy.startswith("Q"):
            continue
        if d.year < ano_inicio or d.month not in (3, 6, 9, 12):
            continue
        q = {3: 1, 6: 2, 9: 3, 12: 4}[d.month]
        per = _normalize_period(d.year, q)
        if per in seen:
            continue
        seen.add(per)
        out.append((c, per))
    return out


def _identify_cols_layout_b(ws, ano_inicio: int, ano_fim: int = 2025) -> list[tuple[int, str]]:
    """Layout CI: row 6 year header '2023A' (somente Q1 col), row 7 'Q1A'/'Q4A'/'2023A'.
    Determina year a partir do header E forca recheck a cada coluna onde year tem 'A'."""
    out: list[tuple[int, str]] = []
    seen: set[str] = set()
    cur_year = None
    cur_year_is_actual = False
    for c in range(2, ws.max_column + 1):
        yr_cell = ws.cell(6, c).value
        if isinstance(yr_cell, str) and len(yr_cell) >= 5:
            try:
                cur_year = int(yr_cell[:4])
                cur_year_is_actual = yr_cell.endswith("A")
            except ValueError:
                pass
        qt = ws.cell(7, c).value
        if not isinstance(qt, str) or not qt.startswith("Q") or "E" in qt:
            continue
        try:
            q = int(qt[1])
        except (ValueError, IndexError):
            continue
        if cur_year is None or not cur_year_is_actual:
            continue
        if cur_year < ano_inicio or cur_year > ano_fim:
            continue
        per = _normalize_period(cur_year, q)
        if per in seen:
            continue
        seen.add(per)
        out.append((c, per))
    return out


def _identify_cols(ws, layout: str, ano_inicio: int) -> list[tuple[int, str]]:
    if layout == "A":
        return _identify_cols_layout_a(ws, ano_inicio)
    if layout == "B":
        return _identify_cols_layout_b(ws, ano_inicio)
    raise ValueError(f"layout desconhecido: {layout}")


def _read_row(ws, cols: list[tuple[int, str]], row: int) -> dict[str, Optional[float]]:
    return {per: _coerce(ws.cell(row, c).value) for c, per in cols}


# ---------------------------------------------------------------------------
# Extractor principal
# ---------------------------------------------------------------------------
def extrair(config: dict, xlsm_path: str, pasta_destino: str,
            ano_inicio: int = 2018) -> list[dict]:
    """
    config keys:
      - layout: "A" ou "B"
      - sheet_is, sheet_bs, sheet_cf: str
      - rows_is, rows_bs, rows_cf: dict[contas_key -> int_row]
      - rows_bs_bpp: dict subset que vai para ITR_bpp
      - signs: dict opcional {contas_key: -1} para inverter sinal
    """
    wb = openpyxl.load_workbook(xlsm_path, data_only=True)
    ws_is = wb[config["sheet_is"]]
    ws_bs = wb[config["sheet_bs"]]
    ws_cf = wb[config["sheet_cf"]]
    layout = config["layout"]

    cols_is = _identify_cols(ws_is, layout, ano_inicio)
    cols_bs = _identify_cols(ws_bs, layout, ano_inicio)
    cols_cf = _identify_cols(ws_cf, layout, ano_inicio)
    print(f"  cols IS={len(cols_is)} BS={len(cols_bs)} CF={len(cols_cf)}")

    signs = config.get("signs", {})

    def _read(ws, cols, rows_map: dict[str, int]) -> dict[str, dict]:
        per_data: dict[str, dict] = {}
        for key, r in rows_map.items():
            if not r:
                continue
            row = _read_row(ws, cols, r)
            sign = signs.get(key, 1)
            for per, v in row.items():
                if v is not None:
                    v = v * sign
                per_data.setdefault(per, {})[key] = v
        return per_data

    is_data = _read(ws_is, cols_is, config["rows_is"])
    bs_data = _read(ws_bs, cols_bs, config["rows_bs"])
    cf_data = _read(ws_cf, cols_cf, config["rows_cf"])

    # Constroi contas_chave (multiplica * 1e6 — XLSMs sao em USD M)
    def m(v):
        return v * 1e6 if v is not None else None

    contas: list[dict] = []
    periodos = sorted(set(is_data) | set(bs_data) | set(cf_data))
    bpp_keys = set(config.get("rows_bs_bpp", {}).keys())

    for periodo in periodos:
        ano = int(periodo[:4])
        ip = is_data.get(periodo, {})
        bp = bs_data.get(periodo, {})
        cp = cf_data.get(periodo, {})

        # DRE
        if ip:
            rec = ip.get("receita_liquida")
            cost = ip.get("custo")
            ebit = ip.get("ebit")
            ebitda = ip.get("ebitda")
            dep = ip.get("depreciacao_amortizacao")
            if dep is None and ebit is not None and ebitda is not None:
                dep = ebitda - ebit
            elif dep is not None:
                dep = abs(dep)
            # Custo proxy se nao vier explicito
            if cost is None and rec is not None and ebitda is not None:
                cost = -(rec - ebitda)
            ni_int = ip.get("despesas_financeiras")  # net interest expense
            if ni_int is not None and ni_int > 0:
                # No XLSM tipico vem positivo (despesa), invertemos
                ni_int = -ni_int
            contas.append({"periodo": periodo, "tipo": "ITR_dre", "ano": ano, "contas": {
                "receita_liquida": m(rec),
                "custo": m(cost) if cost is not None else 0.0,
                "resultado_bruto": m(rec + cost) if (rec is not None and cost is not None) else None,
                "ebit": m(ebit),
                "ebitda": m(ebitda),
                "depreciacao_amortizacao": m(dep),
                "lucro_liquido": m(ip.get("lucro_liquido")),
                "lucro_antes_ir": m(ip.get("lucro_antes_ir")),
                "ir_csll": m(ip.get("ir_csll")),
                "despesas_financeiras": m(ni_int) if ni_int is not None else 0.0,
                "receitas_financeiras": 0.0,
                "resultado_financeiro": m(ni_int),
            }})

        # BPA
        if bp:
            bpa_keys = set(config["rows_bs"].keys()) - bpp_keys
            contas.append({"periodo": periodo, "tipo": "ITR_bpa", "ano": ano, "contas": {
                k: m(bp.get(k)) for k in bpa_keys
            }})
            # BPP
            contas.append({"periodo": periodo, "tipo": "ITR_bpp", "ano": ano, "contas": {
                k: m(bp.get(k)) for k in bpp_keys
            }})

        # DFC
        if cp:
            capex = cp.get("capex")
            if capex is not None and capex > 0:
                capex = -capex
            fco = cp.get("fco")
            contas.append({"periodo": periodo, "tipo": "ITR_dfc", "ano": ano, "contas": {
                "fco": m(fco),
                "capex": m(capex),
                "fci": m(cp.get("fci")),
                "fcf": m((fco + capex)) if (fco is not None and capex is not None) else None,
                "depreciacao_amortizacao": m(abs(cp.get("dep_amort"))) if cp.get("dep_amort") is not None else None,
                "captacao_divida": m(cp.get("captacao_divida")),
                "amortizacao_divida": m(cp.get("amortizacao_divida")),
                "dividendos_pagos": m(cp.get("dividendos_pagos")),
                "buyback": m(cp.get("buyback")),
                "juros_pagos": m(cp.get("juros_pagos")),
            }})

    # Salvar
    destino = Path(pasta_destino)
    destino.mkdir(parents=True, exist_ok=True)
    (destino / "contas_chave.json").write_text(
        json.dumps(contas, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    return contas


# ---------------------------------------------------------------------------
# Configs por empresa
# ---------------------------------------------------------------------------
CONFIG_CI = {
    "layout": "B",
    "sheet_is": "IS",
    "sheet_bs": "BS",
    "sheet_cf": "CF",
    "rows_is": {
        "receita_liquida": 16,           # Total revenues
        "ebitda": 29,                    # EBITA (pre-4Q18 EBITDA)
        "depreciacao_amortizacao": 31,   # Amortization of intangibles
        "ebit": 33,                      # EBIT
        "despesas_financeiras": 34,      # Interest Expense
        "lucro_antes_ir": 37,            # Pre-Tax Income
        "ir_csll": 38,                   # Current Income Taxes
        "lucro_liquido": 51,             # Shareholder's Net Income
    },
    "rows_bs": {
        "caixa": 10,
        "aplicacoes_financeiras_cp": 11,
        "contas_a_receber": 12,
        "estoques_cp": 13,
        "ativo_circulante": 15,
        "imobilizado": 19,
        "intangivel": 20,
        "ativo_total": 23,
        "emprestimos_cp": 29,
        "passivo_circulante": 30,
        "emprestimos_lp": 34,
        "passivo_total": 36,
        "patrimonio_liquido": 43,
    },
    "rows_bs_bpp": {
        "emprestimos_cp": 29,
        "passivo_circulante": 30,
        "emprestimos_lp": 34,
        "passivo_total": 36,
        "patrimonio_liquido": 43,
    },
    "rows_cf": {
        "fco": 27,                       # Cash flow from operations
        "capex": 32,                     # Capital expenditures and other
        "fci": 34,                       # Cash (used in) investing
        "captacao_divida": 38,           # Debt, net (positive=raise)
        "buyback": 39,                   # Repurchases of common stock
        "dividendos_pagos": 40,          # Common dividends paid
        "dep_amort": 12,                 # D&A from CF
    },
    "signs": {
        "despesas_financeiras": -1,      # CI shows positive expense
    },
}


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")

    # CI
    print("=== CI ===")
    contas = extrair(
        CONFIG_CI,
        r"G:\Meu Drive\Análise de Crédito Global\CI\Documentos\861659.CI.N.XLSX",
        r"G:\Meu Drive\Análise de Crédito Global\CI\Dados_EDGAR",
        ano_inicio=2018,
    )
    pers = sorted({c["periodo"] for c in contas})
    print(f"  Periodos: {len(pers)}  | range: {pers[0]} -> {pers[-1]}")
