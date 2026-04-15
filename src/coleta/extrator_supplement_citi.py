"""
Extrator de Financial Data Supplement XLSX do Citigroup (C).

Lê os arquivos XLSX trimestrais via openpyxl e extrai dados estruturados de:
- Summary (ratios, efficiency, ROE)
- Averages - Yields (average balances, interest rates, NIM)
- CET1 Capital / CET1 Capital and TCE (capital ratios, RWA, SLR)
- Deposits (EOP and average deposits)
- ACL RollForward / Allow_ Credit_Losses_Page 1 (NCO, ACL, provisions)
- Non_Accrual Assets (NPA)

Salva em supplement_data.json no mesmo formato que JPM/BAC/BK.
"""

import os
import re
import json
import glob
from datetime import datetime

import openpyxl


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def _parse_number(val) -> float | None:
    """Converte valor de célula em float. Retorna None para vazios/NM."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s in ("—", "–", "―", "-", "NM", "N/M", "n/m"):
        return None
    s = s.replace("$", "").replace(",", "").replace("%", "").strip()
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None


def _quarter_end_date(q: int, year: int) -> str:
    """Retorna data final do trimestre: '2025-12-31'."""
    ends = {1: "03-31", 2: "06-30", 3: "09-30", 4: "12-31"}
    return f"{year}-{ends[q]}"


def _quarter_label(q: int, year: int) -> str:
    """Retorna label: '4Q25'."""
    return f"{q}Q{year % 100:02d}"


def _find_row_by_label(ws, label: str, col: int = 2, start_row: int = 1,
                       end_row: int = 100, partial: bool = True) -> int | None:
    """Encontra número da linha que contém label na coluna col.
    Se partial=True, faz match parcial (case-insensitive)."""
    label_lower = label.lower()
    for row in range(start_row, end_row + 1):
        val = ws.cell(row=row, column=col).value
        if val is None:
            # Also check columns 3, 4, 5 for indented labels
            for c in [3, 4, 5]:
                val = ws.cell(row=row, column=c).value
                if val is not None:
                    break
        if val is None:
            continue
        cell_str = str(val).strip().lower()
        if partial:
            if label_lower in cell_str:
                return row
        else:
            if cell_str == label_lower:
                return row
    return None


def _get_val(ws, row: int, col: int) -> float | None:
    """Pega valor numérico de uma célula."""
    return _parse_number(ws.cell(row=row, column=col).value)


def _detect_quarter_columns_summary(ws) -> list[tuple[int, int, int]]:
    """Detecta colunas de trimestres no Summary/Deposits/Non-Accrual/ACL.

    Busca o padrão Q-label (4Q, 1Q, ...) + year (2024, 2025, ...) em linhas
    consecutivas. Tenta rows 7/8 (Summary) e rows 6/7 (ACL Page 1, Non-Accrual).
    Colunas: 6, 8, 10, 12, 14 (5 trimestres).

    Retorna lista de (quarter, year, col_number).
    """
    # Try multiple row pairs for header detection
    for q_row, yr_row in [(7, 8), (6, 7)]:
        result = []
        for col in [6, 8, 10, 12, 14]:
            q_label = ws.cell(row=q_row, column=col).value
            year_val = ws.cell(row=yr_row, column=col).value
            if q_label is None or year_val is None:
                continue
            q_str = str(q_label).strip()
            m = re.match(r"(\d)Q", q_str)
            if not m:
                continue
            q = int(m.group(1))
            try:
                yr = int(year_val)
            except (ValueError, TypeError):
                continue
            if yr < 100:
                yr += 2000
            result.append((q, yr, col))
        if len(result) >= 3:  # Found enough columns
            return result
    return result  # Return whatever we found
    return result


def _detect_quarter_columns_averages(ws) -> list[tuple[int, int, int]]:
    """Detecta colunas de trimestres no Averages-Yields.

    Row 8 has labels like '4Q24', '3Q25', '4Q25(5)'.
    Average volume columns: 7, 9, 11
    Interest columns: 15, 17, 19
    Rate columns: 22, 24, 26

    Returns list of (quarter, year, vol_col, int_col, rate_col).
    """
    result = []
    for vol_col, int_col, rate_col in [(7, 15, 22), (9, 17, 24), (11, 19, 26)]:
        val = ws.cell(row=8, column=vol_col).value
        if val is None:
            continue
        s = str(val).strip()
        m = re.match(r"(\d)Q(\d{2})", s)
        if not m:
            continue
        q = int(m.group(1))
        yr = 2000 + int(m.group(2))
        result.append((q, yr, vol_col, int_col, rate_col))
    return result


def _detect_quarter_columns_cet1(ws) -> list[tuple[int, int, int]]:
    """Detecta colunas no CET1 Capital sheet.

    Row 8 has month labels (December 31,, March 31,, etc.)
    Row 9 has years (2024, 2025, etc.)
    Columns: 6, 8, 10, 12, 14
    """
    result = []
    month_to_q = {
        "december": 4, "march": 1, "june": 2, "september": 3
    }
    for col in [6, 8, 10, 12, 14]:
        month_label = ws.cell(row=8, column=col).value
        year_val = ws.cell(row=9, column=col).value
        if month_label is None or year_val is None:
            continue
        month_str = str(month_label).strip().lower()
        q = None
        for m_key, m_q in month_to_q.items():
            if m_key in month_str:
                q = m_q
                break
        if q is None:
            continue
        yr_str = str(year_val).strip()
        # Remove footnote markers like "(2)", "(5)", etc.
        yr_str = re.sub(r"\(.*?\)", "", yr_str).strip()
        m_yr = re.search(r"(\d{4})", yr_str)
        if not m_yr:
            continue
        yr = int(m_yr.group(1))
        if yr < 100:
            yr += 2000
        result.append((q, yr, col))
    return result


def _detect_quarter_columns_acl_rollforward(ws) -> list[tuple[int, int, int]]:
    """Detecta colunas de trimestres no ACL RollForward.

    Row 8 has build/release labels: '1Q25', '2Q25', '3Q25', '4Q25'
    The 'Balance' column at the end (col 34) has the end-of-period balance.
    Balance at beginning is col 20 (for the year start) or col 6 (prior year start).

    For our purposes, we want EOP ACLL and NCO per quarter, which come from
    Allow_ Credit_Losses_Page 1 instead (simpler quarterly layout).
    """
    result = []
    for col in [6, 8, 10, 12, 14]:
        val = ws.cell(row=7, column=col).value
        if val is None:
            val = ws.cell(row=8, column=col).value
        if val is None:
            continue
        s = str(val).strip()
        m = re.match(r"(\d)Q(\d{2})", s)
        if m:
            q = int(m.group(1))
            yr = 2000 + int(m.group(2))
            result.append((q, yr, col))
    return result


# ---------------------------------------------------------------------------
# Parsers por sheet
# ---------------------------------------------------------------------------

def _parse_summary(ws, q: int, yr: int, col: int) -> dict:
    """Extrai dados do Summary para um trimestre."""
    data = {}

    # Regulatory capital ratios
    row = _find_row_by_label(ws, "Common Equity Tier 1 (CET1) Capital ratio")
    if row:
        data["cet1_ratio_summary"] = _get_val(ws, row, col)

    row = _find_row_by_label(ws, "Tier 1 Capital ratio")
    if row:
        data["tier1_ratio"] = _get_val(ws, row, col)

    row = _find_row_by_label(ws, "Total Capital ratio")
    if row:
        data["total_capital_ratio_summary"] = _get_val(ws, row, col)

    row = _find_row_by_label(ws, "Supplementary Leverage ratio")
    if row:
        data["slr_summary"] = _get_val(ws, row, col)

    row = _find_row_by_label(ws, "Efficiency ratio")
    if row:
        data["efficiency_ratio"] = _get_val(ws, row, col)

    row = _find_row_by_label(ws, "Return on average assets")
    if row:
        data["roa"] = _get_val(ws, row, col)

    row = _find_row_by_label(ws, "Return on tangible common equity")
    if row:
        data["rotce"] = _get_val(ws, row, col)

    # Net credit losses from Summary
    row = _find_row_by_label(ws, "Net credit losses (NCLs)")
    if row:
        data["ncl_summary"] = _get_val(ws, row, col)

    # Balance sheet (in billions)
    row = _find_row_by_label(ws, "Total average assets", start_row=55)
    if row:
        v = _get_val(ws, row, col)
        if v is not None:
            data["total_avg_assets_bn"] = v

    row = _find_row_by_label(ws, "Total loans", start_row=55)
    if row:
        v = _get_val(ws, row, col)
        if v is not None:
            data["total_loans_eop_bn"] = v

    row = _find_row_by_label(ws, "Total deposits", start_row=55)
    if row:
        v = _get_val(ws, row, col)
        if v is not None:
            data["total_deposits_eop_bn"] = v

    return data


def _parse_averages_yields(ws, q: int, yr: int, vol_col: int,
                           int_col: int, rate_col: int) -> dict:
    """Extrai dados de Average Balances e Yields para um trimestre."""
    ab = {}
    yr_data = {}

    # --- ASSETS ---
    row = _find_row_by_label(ws, "Total average interest-earning assets")
    if row:
        ab["avg_earning_assets"] = _get_val(ws, row, vol_col)
        yr_data["avg_earning_assets_yield"] = _get_val(ws, row, rate_col)

    row = _find_row_by_label(ws, "Total loans (net of unearned income)")
    if row:
        ab["avg_loans"] = _get_val(ws, row, vol_col)
        yr_data["avg_loans_yield"] = _get_val(ws, row, rate_col)

    # --- LIABILITIES ---
    row = _find_row_by_label(ws, "Deposits", start_row=20, end_row=25)
    if row:
        ab["avg_ib_deposits"] = _get_val(ws, row, vol_col)
        yr_data["avg_ib_deposits_rate"] = _get_val(ws, row, rate_col)

    row = _find_row_by_label(ws, "Total average interest-bearing liabilities")
    if row:
        ab["avg_total_ib_liabilities"] = _get_val(ws, row, vol_col)
        yr_data["avg_total_ib_liabilities_rate"] = _get_val(ws, row, rate_col)

    # --- NIM ---
    row = _find_row_by_label(ws, "Net interest income as a % of average interest-earning assets")
    if row:
        yr_data["nim"] = _get_val(ws, row, rate_col)
        nii_val = _get_val(ws, row, int_col)
        if nii_val is not None:
            yr_data["nii_fte"] = nii_val

    # Interest spread = earning assets yield - IB liabilities rate
    ea_yld = yr_data.get("avg_earning_assets_yield")
    ib_rate = yr_data.get("avg_total_ib_liabilities_rate")
    if ea_yld is not None and ib_rate is not None:
        yr_data["interest_spread"] = round(ea_yld - ib_rate, 6)

    return {"avg_balances": ab, "yields_rates": yr_data}


def _parse_cet1_capital(ws, q: int, yr: int, col: int) -> dict:
    """Extrai dados do CET1 Capital sheet."""
    cap = {}

    row = _find_row_by_label(ws, "CET1 Capital ratio")
    if row:
        cap["cet1_ratio"] = _get_val(ws, row, col)

    # Find CET1 Capital line (exact "CET1 Capital" or "Common Equity Tier 1 Capital (CET1)")
    for r in range(10, 35):
        val = ws.cell(row=r, column=2).value
        if val is None:
            continue
        label = str(val).strip()
        if label in ("CET1 Capital",
                      "Common Equity Tier 1 Capital (CET1)"):
            cap["cet1_capital"] = _get_val(ws, r, col)
            break

    row = _find_row_by_label(ws, "Risk-Weighted Assets")
    if row:
        cap["rwa_standardized"] = _get_val(ws, row, col)

    row = _find_row_by_label(ws, "Supplementary Leverage ratio")
    if row:
        cap["slr"] = _get_val(ws, row, col)

    row = _find_row_by_label(ws, "Total Tier 1 Capital")
    if row:
        cap["tier1_capital"] = _get_val(ws, row, col)

    row = _find_row_by_label(ws, "Total Leverage Exposure")
    if row:
        cap["total_leverage_exposure"] = _get_val(ws, row, col)

    # Total Capital ratio comes from Summary, not here
    # LCR / NSFR - preenchido depois via _extract_lcr_from_10q
    cap["lcr"] = None
    cap["nsfr"] = None

    return cap


def _parse_deposits(ws, q: int, yr: int, col: int) -> dict:
    """Extrai dados do Deposits sheet."""
    dep = {}

    row = _find_row_by_label(ws, "Total deposits", start_row=35, end_row=45)
    if row is None:
        row = _find_row_by_label(ws, "Total deposits")
    if row:
        # Check if it's EOP or average
        label = str(ws.cell(row=row, column=2).value or "").strip().lower()
        if "eop" in label:
            v = _get_val(ws, row, col)
            if v is not None:
                dep["total_deposits_eop"] = v  # in billions

    row = _find_row_by_label(ws, "Total deposits", start_row=39, end_row=45)
    if row:
        label = str(ws.cell(row=row, column=2).value or "").strip().lower()
        if "average" in label:
            v = _get_val(ws, row, col)
            if v is not None:
                dep["avg_total_deposits"] = v  # in billions

    return dep


def _parse_credit_quality_acl(ws_acl_p1, q: int, yr: int, col: int) -> dict:
    """Extrai NCO e ACLL do Allow_ Credit_Losses_Page 1."""
    cq = {}

    # NCLs (Net credit losses)
    row = _find_row_by_label(ws_acl_p1, "Net credit (losses) / recoveries on loans")
    if row:
        v = _get_val(ws_acl_p1, row, col)
        if v is not None:
            cq["nco_total"] = abs(v)  # Store as positive

    # ACLL at end of period (a)
    row = _find_row_by_label(ws_acl_p1, "ACLL at end of period")
    if row:
        cq["acl_loans"] = _get_val(ws_acl_p1, row, col)

    # Total ACLL as pct of total loans
    row = _find_row_by_label(ws_acl_p1, "Total ACLL as a percentage of total loans")
    if row:
        cq["acl_pct_loans"] = _get_val(ws_acl_p1, row, col)

    # Total ACL (ACLL + ACLUC) - label spans 2 rows, data on second row
    row = _find_row_by_label(ws_acl_p1, "Total allowance for credit losses on loans")
    if row:
        v = _get_val(ws_acl_p1, row, col)
        if v is None:
            # Data might be on the next row (multi-line label)
            v = _get_val(ws_acl_p1, row + 1, col)
        cq["acl_total"] = v

    return cq


def _parse_non_accrual(ws, q: int, yr: int, col: int) -> dict:
    """Extrai NPA do Non_Accrual Assets sheet."""
    npa = {}

    row = _find_row_by_label(ws, "Total non-accrual loans (NAL)")
    if row:
        npa["npa"] = _get_val(ws, row, col)

    row = _find_row_by_label(ws, "NAL as a percentage of total loans")
    if row:
        npa["nal_pct_loans"] = _get_val(ws, row, col)

    return npa


# ---------------------------------------------------------------------------
# Processamento de um arquivo XLSX
# ---------------------------------------------------------------------------

def _get_sheet(wb, *names):
    """Tenta encontrar sheet por nomes alternativos."""
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
        # Try with trailing space
        if name + " " in wb.sheetnames:
            return wb[name + " "]
    return None


def _process_file(xlsx_path: str) -> list[dict]:
    """Processa um arquivo XLSX e retorna lista de dicts por trimestre."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    fname = os.path.basename(xlsx_path)
    results = {}  # key = (q, yr) -> dict

    # --- Summary (5 quarters) ---
    ws_sum = _get_sheet(wb, "Summary")
    if ws_sum:
        qcols = _detect_quarter_columns_summary(ws_sum)
        for q, yr, col in qcols:
            key = (q, yr)
            if key not in results:
                results[key] = _init_entry(q, yr, fname)
            data = _parse_summary(ws_sum, q, yr, col)
            results[key]["_summary"] = data

    # --- Averages - Yields (3 quarters) ---
    ws_ay = _get_sheet(wb, "Averages - Yields", "Averages - Yields ")
    if ws_ay:
        ay_cols = _detect_quarter_columns_averages(ws_ay)
        for q, yr, vol_col, int_col, rate_col in ay_cols:
            key = (q, yr)
            if key not in results:
                results[key] = _init_entry(q, yr, fname)
            data = _parse_averages_yields(ws_ay, q, yr, vol_col, int_col, rate_col)
            results[key]["avg_balances"] = data["avg_balances"]
            results[key]["yields_rates"] = data["yields_rates"]

    # --- CET1 Capital (5 quarters) ---
    ws_cet1 = _get_sheet(wb, "CET1 Capital", "CET1 Capital and TCE")
    if ws_cet1:
        cet1_cols = _detect_quarter_columns_cet1(ws_cet1)
        for q, yr, col in cet1_cols:
            key = (q, yr)
            if key not in results:
                results[key] = _init_entry(q, yr, fname)
            results[key]["capital"] = _parse_cet1_capital(ws_cet1, q, yr, col)

    # --- Deposits (5 quarters) ---
    ws_dep = _get_sheet(wb, "Deposits")
    if ws_dep:
        dep_cols = _detect_quarter_columns_summary(ws_dep)  # Same header format
        for q, yr, col in dep_cols:
            key = (q, yr)
            if key not in results:
                results[key] = _init_entry(q, yr, fname)
            results[key]["_deposits"] = _parse_deposits(ws_dep, q, yr, col)

    # --- Allow_ Credit_Losses_Page 1 (5 quarters for NCO/ACLL) ---
    ws_acl = _get_sheet(wb, "Allow_ Credit_Losses_Page 1")
    if ws_acl:
        acl_cols = _detect_quarter_columns_summary(ws_acl)
        for q, yr, col in acl_cols:
            key = (q, yr)
            if key not in results:
                results[key] = _init_entry(q, yr, fname)
            results[key]["_credit_acl"] = _parse_credit_quality_acl(
                ws_acl, q, yr, col
            )

    # --- Non_Accrual Assets (5 quarters) ---
    ws_na = _get_sheet(wb, "Non_Accrual Assets")
    if ws_na:
        na_cols = _detect_quarter_columns_summary(ws_na)
        for q, yr, col in na_cols:
            key = (q, yr)
            if key not in results:
                results[key] = _init_entry(q, yr, fname)
            results[key]["_npa"] = _parse_non_accrual(ws_na, q, yr, col)

    wb.close()
    return list(results.values())


def _init_entry(q: int, yr: int, fname: str) -> dict:
    """Cria entrada padrão para um trimestre."""
    return {
        "periodo": _quarter_end_date(q, yr),
        "trimestre": _quarter_label(q, yr),
        "fonte": fname,
        "avg_balances": {},
        "yields_rates": {},
        "capital": {},
        "_summary": {},
        "_deposits": {},
        "_credit_acl": {},
        "_npa": {},
    }


# ---------------------------------------------------------------------------
# Consolidação
# ---------------------------------------------------------------------------

def _merge_quarter_data(entries: list[dict]) -> dict:
    """Merge multiple entries for the same quarter (from different files).
    Later files (newer) overwrite older ones, but only fill empty fields."""
    if not entries:
        return {}
    # Start with the first, overlay with subsequent
    merged = entries[0].copy()
    for e in entries[1:]:
        for section in ["avg_balances", "yields_rates", "capital",
                        "_summary", "_deposits", "_credit_acl", "_npa"]:
            if section not in merged:
                merged[section] = {}
            src = e.get(section, {})
            for k, v in src.items():
                if v is not None and (k not in merged[section] or merged[section][k] is None):
                    merged[section][k] = v
        # Update fonte to most recent
        merged["fonte"] = e.get("fonte", merged.get("fonte"))
    return merged


def _finalize_entry(entry: dict) -> dict:
    """Consolida dados internos no formato final."""
    summary = entry.get("_summary", {})
    deposits = entry.get("_deposits", {})
    credit_acl = entry.get("_credit_acl", {})
    npa_data = entry.get("_npa", {})
    capital = entry.get("capital", {})
    ab = entry.get("avg_balances", {})
    yr = entry.get("yields_rates", {})

    # --- avg_balances ---
    # avg_total_assets from Summary (in billions -> convert to millions)
    if "avg_total_assets" not in ab or ab.get("avg_total_assets") is None:
        v = summary.get("total_avg_assets_bn")
        if v is not None:
            ab["avg_total_assets"] = round(v * 1000, 1)

    # avg_total_deposits from Deposits sheet (in billions -> millions)
    if "avg_total_deposits" not in ab or ab.get("avg_total_deposits") is None:
        v = deposits.get("avg_total_deposits")
        if v is not None:
            ab["avg_total_deposits"] = round(v * 1000, 1)

    # Citi doesn't split IB/NIB in the supplement; avg_ib_deposits from
    # Averages-Yields "Deposits" line (which is interest-bearing deposits)
    # NIB = total - IB
    ib = ab.get("avg_ib_deposits")
    total_dep = ab.get("avg_total_deposits")
    if ib is not None and total_dep is not None:
        ab["avg_nib_deposits"] = round(total_dep - ib, 1)

    # --- capital ---
    # Merge summary ratios into capital if not already set
    if capital.get("cet1_ratio") is None:
        capital["cet1_ratio"] = summary.get("cet1_ratio_summary")
    if capital.get("slr") is None:
        capital["slr"] = summary.get("slr_summary")
    capital["total_capital_ratio"] = summary.get("total_capital_ratio_summary",
                                                  capital.get("total_capital_ratio"))

    # --- credit_quality ---
    cq = {}
    cq["nco_total"] = credit_acl.get("nco_total", summary.get("ncl_summary"))
    cq["acl_loans"] = credit_acl.get("acl_loans")
    cq["acl_pct_loans"] = credit_acl.get("acl_pct_loans")
    cq["acl_total"] = credit_acl.get("acl_total")
    cq["npa"] = npa_data.get("npa")

    # Build final
    final = {
        "periodo": entry["periodo"],
        "trimestre": entry["trimestre"],
        "fonte": entry["fonte"],
        "avg_balances": ab,
        "yields_rates": yr,
        "capital": capital,
        "credit_quality": cq,
    }
    return final


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------

def extrair_supplement_citi(pasta_docs: str, pasta_destino: str) -> list[dict]:
    """Extrai dados de todos os Financial Data Supplement XLSX do Citigroup.

    Args:
        pasta_docs: Caminho para pasta com os XLSX (financial-supplement-*.xlsx)
        pasta_destino: Caminho para pasta de destino do JSON

    Returns:
        Lista de dicionários com dados extraídos por trimestre.
    """
    xlsx_files = sorted(glob.glob(os.path.join(pasta_docs, "financial-supplement-*.xlsx")))
    if not xlsx_files:
        print(f"[WARN] Nenhum XLSX encontrado em {pasta_docs}")
        return []

    # Collect all quarter data from all files
    all_quarters = {}  # (q, yr) -> list of entries

    for xlsx_path in xlsx_files:
        fname = os.path.basename(xlsx_path)
        print(f"  Processando {fname}...")

        try:
            entries = _process_file(xlsx_path)
        except Exception as e:
            print(f"    [ERRO] Falha ao processar: {e}")
            import traceback
            traceback.print_exc()
            continue

        for entry in entries:
            m = re.match(r"(\d)Q(\d{2})", entry["trimestre"])
            if not m:
                continue
            q = int(m.group(1))
            yr = 2000 + int(m.group(2))
            key = (q, yr)
            if key not in all_quarters:
                all_quarters[key] = []
            all_quarters[key].append(entry)

    # Merge & finalize
    resultados = []
    for key in sorted(all_quarters.keys(), key=lambda k: (k[1], k[0])):
        entries = all_quarters[key]
        merged = _merge_quarter_data(entries)
        final = _finalize_entry(merged)
        resultados.append(final)

    # Enriquecer com LCR dos 10-Q/10-K filings da EDGAR
    periodos = [r["periodo"] for r in resultados]
    print("  Extraindo LCR dos 10-Q/10-K filings...")
    from .extrator_supplement_bac import _extract_lcr_from_10q
    lcr_map = _extract_lcr_from_10q("C", periodos)
    for entry in resultados:
        per = entry.get("periodo")
        if per in lcr_map:
            cap = entry.get("capital", {})
            cap["lcr"] = lcr_map[per]
            entry["capital"] = cap

    # Save
    os.makedirs(pasta_destino, exist_ok=True)
    out_path = os.path.join(pasta_destino, "supplement_data.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(resultados, f, indent=2, ensure_ascii=False)

    print(f"\n  Salvo {len(resultados)} trimestres em {out_path}")
    return resultados


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    pasta_docs = (sys.argv[1] if len(sys.argv) > 1
                  else "G:/Meu Drive/Análise de Crédito Financeiras/C/Documentos")
    pasta_destino = (sys.argv[2] if len(sys.argv) > 2
                     else "G:/Meu Drive/Análise de Crédito Financeiras/C/Dados_EDGAR")

    dados = extrair_supplement_citi(pasta_docs, pasta_destino)

    if dados:
        print("\n=== Últimos 2 trimestres ===")
        for entry in dados[-2:]:
            print(json.dumps(entry, indent=2, ensure_ascii=False))
