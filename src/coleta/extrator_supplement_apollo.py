"""
Extrator de Financial Supplement XLSX da Apollo Global Management (APO).

Le os arquivos financial-supplement-XQ2Y.xlsx e extrai metricas trimestrais
de Asset Manager, salvando como am_data.json.

Estrategia:
- Usa o arquivo 4Q25 (12 trimestres: 1Q23-4Q25)
- Usa o arquivo 4Q23 (12 trimestres: 1Q21-4Q23, pegamos 1Q22-4Q22)
- Combina para cobrir 1Q22 ate 4Q25 (16 trimestres)
"""

import json
import os
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Mapeamento de header de coluna -> data fim do trimestre
# ---------------------------------------------------------------------------
def _quarter_label_to_date(label: str) -> str | None:
    """Converte '1Q\\'23' ou '1Q23' -> '2023-03-31'."""
    if not label or not isinstance(label, str):
        return None
    label = label.strip().replace("'", "").replace("\u2019", "")
    # Formatos: "1Q23", "2Q23", etc.
    import re
    m = re.match(r"(\d)Q(\d{2,4})", label)
    if not m:
        return None
    q = int(m.group(1))
    y = int(m.group(2))
    if y < 100:
        y += 2000
    end_months = {1: "03-31", 2: "06-30", 3: "09-30", 4: "12-31"}
    if q not in end_months:
        return None
    return f"{y}-{end_months[q]}"


def _quarter_label_to_tri(label: str) -> str | None:
    """Converte '1Q\\'23' -> 'Q1/23'."""
    if not label or not isinstance(label, str):
        return None
    label = label.strip().replace("'", "").replace("\u2019", "")
    import re
    m = re.match(r"(\d)Q(\d{2,4})", label)
    if not m:
        return None
    q = m.group(1)
    y = m.group(2)
    if len(y) == 4:
        y = y[2:]
    return f"Q{q}/{y}"


# ---------------------------------------------------------------------------
# Helpers para ler dados de uma sheet
# ---------------------------------------------------------------------------
def _build_col_map(ws, header_row: int = 3) -> dict[int, str]:
    """
    Retorna {col_index: 'YYYY-MM-DD'} para as colunas trimestrais.
    Ignora colunas FY.
    """
    col_map = {}
    for cell in ws[header_row]:
        if cell.value and isinstance(cell.value, str):
            val = cell.value.strip()
            if val.startswith("FY"):
                continue
            dt = _quarter_label_to_date(val)
            if dt:
                col_map[cell.column] = dt
    return col_map


def _find_row(ws, search_text: str, exact: bool = False, starts: bool = False, max_row: int = 70) -> int | None:
    """Encontra a linha cujo label (col A) contem search_text.

    exact: label (sem footnotes) deve ser igual ao search_text
    starts: label deve comecar com search_text (case-insensitive)
    default: label contem search_text (case-insensitive)
    """
    import re
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=1):
        cell = row[0]
        if cell.value and isinstance(cell.value, str):
            label = cell.value.strip()
            if exact:
                clean = re.sub(r'[\d,\s]+$', '', label).strip()
                if clean == search_text:
                    return cell.row
            elif starts:
                if label.lower().startswith(search_text.lower()):
                    return cell.row
            else:
                if search_text.lower() in label.lower():
                    return cell.row
    return None


def _get_row_data(ws, row_num: int, col_map: dict[int, str]) -> dict[str, float | None]:
    """Retorna {periodo: valor} para uma linha."""
    data = {}
    for col_idx, periodo in col_map.items():
        cell = ws.cell(row=row_num, column=col_idx)
        val = cell.value
        if val is not None and isinstance(val, (int, float)):
            data[periodo] = val
        else:
            data[periodo] = None
    return data


# ---------------------------------------------------------------------------
# Extracao de um arquivo XLSX
# ---------------------------------------------------------------------------
def _extrair_de_arquivo(filepath: str, periodos_filtro: set[str] | None = None) -> dict[str, dict]:
    """
    Extrai dados de um arquivo de financial supplement.
    Retorna {periodo: {campo: valor, ...}}.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    resultado = {}

    # ---- SUMMARY ----
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        col_map = _build_col_map(ws)

        mapa_summary = {
            "fre": "Fee Related Earnings",
            "sre": "Spread Related Earnings",
            "fsre": "Fee and Spread Related Earnings",
            "pii": "Principal Investing Income",
            "segment_income": "Segment Income",
            "ani": "Adjusted Net Income",
        }
        # Per-share: labels podem variar (maiuscula/minuscula "per share"/"per Share")
        mapa_per_share = {
            "fre_per_share": "FRE per",
            "sre_per_share": "SRE per",
            "ani_per_share": "ANI per",
        }

        for campo, search in mapa_summary.items():
            row_num = _find_row(ws, search)
            if row_num:
                row_data = _get_row_data(ws, row_num, col_map)
                for periodo, val in row_data.items():
                    if periodos_filtro and periodo not in periodos_filtro:
                        continue
                    resultado.setdefault(periodo, {})[campo] = val

        for campo, search in mapa_per_share.items():
            row_num = _find_row(ws, search)
            if row_num:
                row_data = _get_row_data(ws, row_num, col_map)
                for periodo, val in row_data.items():
                    if periodos_filtro and periodo not in periodos_filtro:
                        continue
                    resultado.setdefault(periodo, {})[campo] = val

    # ---- TOTAL SEGMENT EARNINGS ----
    if "Total Segment Earnings" in wb.sheetnames:
        ws = wb["Total Segment Earnings"]
        col_map = _build_col_map(ws)

        mapa_tse = {
            "mgmt_fees": "Total management fees",
            "capital_solutions_fees": "Capital solutions fees",
            "fee_related_perf_fees": "Fee related performance fees",
            "fee_related_revenues": "Fee Related Revenues",
            "net_investment_spread": "Net investment spread",
            "perf_fees_realized": "Realized performance fees",
        }

        for campo, search in mapa_tse.items():
            row_num = _find_row(ws, search)
            if row_num:
                row_data = _get_row_data(ws, row_num, col_map)
                for periodo, val in row_data.items():
                    if periodos_filtro and periodo not in periodos_filtro:
                        continue
                    resultado.setdefault(periodo, {})[campo] = val

    # ---- RECONCILIATION_NIA AND ALTS ----
    recon_sheet = None
    for name in ["Reconciliation_NIA and Alts", "Reconciliation_NIA"]:
        if name in wb.sheetnames:
            recon_sheet = name
            break

    if recon_sheet:
        ws = wb[recon_sheet]
        col_map = _build_col_map(ws)

        # Total investments
        row_num = _find_row(ws, "Total investments")
        if row_num:
            row_data = _get_row_data(ws, row_num, col_map)
            for periodo, val in row_data.items():
                if periodos_filtro and periodo not in periodos_filtro:
                    continue
                resultado.setdefault(periodo, {})["total_investments"] = val

        # Gross invested assets (starts= to avoid matching "Total adjustments to arrive at gross invested assets")
        row_num = _find_row(ws, "Gross invested assets", starts=True)
        if row_num:
            row_data = _get_row_data(ws, row_num, col_map)
            for periodo, val in row_data.items():
                if periodos_filtro and periodo not in periodos_filtro:
                    continue
                resultado.setdefault(periodo, {})["gross_invested_assets"] = val

        # Net invested assets (starts= to avoid matching other rows)
        row_num = _find_row(ws, "Net invested assets", starts=True)
        if row_num:
            row_data = _get_row_data(ws, row_num, col_map)
            for periodo, val in row_data.items():
                if periodos_filtro and periodo not in periodos_filtro:
                    continue
                resultado.setdefault(periodo, {})["net_invested_assets"] = val

    # ---- RS FLOWS AND IA ----
    rs_sheet = None
    for name in ["RS Flows and IA", "RS Flows and NIA"]:
        if name in wb.sheetnames:
            rs_sheet = name
            break

    if rs_sheet:
        ws = wb[rs_sheet]
        col_map = _build_col_map(ws)

        mapa_rs = {
            "rs_gross_organic_inflows": "Gross organic inflows",
            "rs_total_gross_inflows": "Total gross inflows",
            "rs_gross_outflows": "Gross outflows",
            "rs_net_flows": "Net flows",
        }

        for campo, search in mapa_rs.items():
            row_num = _find_row(ws, search)
            if row_num:
                row_data = _get_row_data(ws, row_num, col_map)
                for periodo, val in row_data.items():
                    if periodos_filtro and periodo not in periodos_filtro:
                        continue
                    resultado.setdefault(periodo, {})[campo] = val

        # Invested assets section
        for search, campo in [
            ("Gross invested assets", "rs_gross_invested_assets"),
            ("Net invested assets", "rs_net_invested_assets"),
        ]:
            row_num = _find_row(ws, search, starts=True)
            if row_num:
                # The invested assets section may have its own header row
                # Try using the same col_map first, then check for a second header
                row_data = _get_row_data(ws, row_num, col_map)
                for periodo, val in row_data.items():
                    if periodos_filtro and periodo not in periodos_filtro:
                        continue
                    resultado.setdefault(periodo, {})[campo] = val

    wb.close()
    return resultado


# ---------------------------------------------------------------------------
# Funcao principal
# ---------------------------------------------------------------------------
def extrair_supplement_apollo(pasta_docs: str, pasta_destino: str) -> dict:
    """
    Extrai dados dos Financial Supplements da Apollo e salva am_data.json.

    Estrategia:
    - 4Q25: cobre 1Q23-4Q25 (12 trimestres)
    - 4Q23: cobre 1Q22-4Q22 (pega trimestres nao cobertos pelo 4Q25)

    Args:
        pasta_docs: Pasta com os arquivos XLSX
        pasta_destino: Pasta para salvar am_data.json

    Returns:
        dict com os dados extraidos
    """
    pasta_docs = str(pasta_docs)
    pasta_destino = str(pasta_destino)

    # Arquivos principais
    arquivo_4q25 = os.path.join(pasta_docs, "financial-supplement-4Q25.xlsx")
    arquivo_4q23 = os.path.join(pasta_docs, "financial-supplement-4Q23.xlsx")

    if not os.path.exists(arquivo_4q25):
        raise FileNotFoundError(f"Arquivo nao encontrado: {arquivo_4q25}")

    print("[APO_SUPPLEMENT] Extraindo dados de 4Q25 (1Q23-4Q25)...")
    dados_4q25 = _extrair_de_arquivo(arquivo_4q25)

    # Periodos a buscar em 4Q23: 1Q22 ate 4Q22
    periodos_22 = {
        "2022-03-31", "2022-06-30", "2022-09-30", "2022-12-31",
    }

    dados_4q23 = {}
    if os.path.exists(arquivo_4q23):
        print("[APO_SUPPLEMENT] Extraindo dados de 4Q23 (1Q22-4Q22)...")
        dados_4q23 = _extrair_de_arquivo(arquivo_4q23, periodos_filtro=periodos_22)
    else:
        print(f"[APO_SUPPLEMENT] AVISO: {arquivo_4q23} nao encontrado, sem dados 2022")

    # Combinar: 4Q23 (para 2022) + 4Q25 (para 2023-2025)
    todos = {}
    todos.update(dados_4q23)
    todos.update(dados_4q25)

    # Montar periodos no formato esperado
    periodos_json = []
    for periodo in sorted(todos.keys()):
        d = todos[periodo]

        fre = d.get("fre")
        frr = d.get("fee_related_revenues")

        # FRE margin = FRE / Fee Related Revenues
        fre_margin_pct = None
        if fre is not None and frr is not None and frr != 0:
            fre_margin_pct = round((fre / frr) * 100, 1)

        # DE (distributable earnings) = Segment Income para Apollo
        de = d.get("segment_income")

        # Total AUM: Apollo reports gross invested assets as proxy for total AUM
        # (RS segment). The "real" total AUM is reported elsewhere but
        # gross_invested_assets from the reconciliation/RS sheets is a good proxy.
        total_aum = d.get("rs_gross_invested_assets") or d.get("gross_invested_assets")

        entry = {
            "periodo": periodo,
            "trimestre": _quarter_label_to_tri_from_date(periodo),
            "fre": fre,
            "fre_margin_pct": fre_margin_pct,
            "sre": d.get("sre"),
            "de": de,
            "ani": d.get("ani"),
            "fsre": d.get("fsre"),
            "pii": d.get("pii"),
            "total_aum": total_aum,
            "fee_paying_aum": None,
            "dry_powder": None,
            "permanent_capital_pct": None,
            "net_accrued_performance": None,
            "management_fees": d.get("mgmt_fees"),
            "advisory_fees": None,
            "performance_fees_realized": d.get("perf_fees_realized"),
            "performance_fees_unrealized": None,
            "capital_solutions_fees": d.get("capital_solutions_fees"),
            "fee_related_perf_fees": d.get("fee_related_perf_fees"),
            "fee_related_revenues": frr,
            "net_investment_spread": d.get("net_investment_spread"),
            "segment_income": d.get("segment_income"),
            "fre_per_share": d.get("fre_per_share"),
            "sre_per_share": d.get("sre_per_share"),
            "ani_per_share": d.get("ani_per_share"),
            "gross_invested_assets": d.get("gross_invested_assets") or d.get("rs_gross_invested_assets"),
            "net_invested_assets": d.get("net_invested_assets") or d.get("rs_net_invested_assets"),
            "rs_gross_organic_inflows": d.get("rs_gross_organic_inflows"),
            "rs_total_gross_inflows": d.get("rs_total_gross_inflows"),
            "rs_net_flows": d.get("rs_net_flows"),
        }

        periodos_json.append(entry)

    resultado = {
        "ticker": "APO",
        "fonte": "Financial Supplement XLSX (parser direto)",
        "periodos": periodos_json,
    }

    # Salvar
    os.makedirs(pasta_destino, exist_ok=True)
    output_path = os.path.join(pasta_destino, "am_data.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    print(f"[APO_SUPPLEMENT] Salvo: {output_path} ({len(periodos_json)} periodos)")
    return resultado


def _quarter_label_to_tri_from_date(date_str: str) -> str:
    """Converte '2023-03-31' -> 'Q1/23'."""
    parts = date_str.split("-")
    y = int(parts[0])
    m = int(parts[1])
    q = (m - 1) // 3 + 1
    return f"Q{q}/{y % 100:02d}"


# ---------------------------------------------------------------------------
# Execucao direta
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    PASTA_DOCS = "G:/Meu Drive/Analise de Credito Financeiras/APO/Documentos"
    PASTA_DESTINO = "G:/Meu Drive/Analise de Credito Financeiras/APO/Dados_Extraidos"

    # Tentar path alternativo com acentos
    if not os.path.isdir(PASTA_DOCS):
        PASTA_DOCS = "G:/Meu Drive/Análise de Crédito Financeiras/APO/Documentos"
        PASTA_DESTINO = "G:/Meu Drive/Análise de Crédito Financeiras/APO/Dados_Extraidos"

    dados = extrair_supplement_apollo(PASTA_DOCS, PASTA_DESTINO)

    # Imprimir ultimos 4 trimestres
    print("\n" + "=" * 80)
    print("ULTIMOS 4 TRIMESTRES:")
    print("=" * 80)
    for p in dados["periodos"][-4:]:
        print(f"\n--- {p['trimestre']} ({p['periodo']}) ---")
        for k, v in p.items():
            if k in ("periodo", "trimestre"):
                continue
            if v is not None:
                if isinstance(v, float) and abs(v) < 10:
                    print(f"  {k:30s}: {v}")
                else:
                    print(f"  {k:30s}: {v:,.0f}" if isinstance(v, (int, float)) else f"  {k:30s}: {v}")
