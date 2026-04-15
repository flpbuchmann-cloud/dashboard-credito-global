"""
Extrator de Earnings Supplement XLSX do JPMorgan Chase (JPM).

Lê os XLSX trimestrais e extrai dados estruturados de:
- Average Balances & Yields (Page 6)
- Capital & Balance Sheet (Pages 2, 9)
- Credit Quality (Pages 25, 26, 27)

Salva em supplement_data.json.
"""

import os
import re
import json
import glob
import warnings
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def _parse_number(val) -> float | None:
    """Converte valor de célula (int, float ou str) em float.
    Retorna None se não parsear."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        text = val.strip().replace("$", "").replace("%", "").replace(",", "").strip()
        # Remove footnote markers
        text = re.sub(r'\([a-z]\)', '', text).strip()
        text = re.sub(r'\s+', '', text)
        if not text or text in ("—", "–", "―", "�", "NM"):
            return None
        negative = False
        if text.startswith("(") and text.endswith(")"):
            negative = True
            text = text[1:-1].strip()
        try:
            v = float(text)
            return -v if negative else v
        except ValueError:
            return None
    return None


def _pct_to_decimal(v: float | None) -> float | None:
    """Converte valor de taxa/yield para decimal.
    Detecta se o valor já está em formato decimal (< 1) ou em percentual (>= 1).
    Ex: 2.61 (percentual) -> 0.0261;  0.0261 (já decimal) -> 0.0261."""
    if v is None:
        return None
    if abs(v) < 1:
        # Já está em formato decimal (ex: 0.0261 = 2.61%)
        return round(v, 6)
    # Formato percentual (ex: 2.61 = 2.61%)
    return round(v / 100, 6)


def _quarter_to_date(quarter_str: str) -> str:
    """Converte '4Q25' -> '2025-12-31', etc."""
    m = re.match(r"(\d)Q(\d{2})", quarter_str)
    if not m:
        return quarter_str
    q = int(m.group(1))
    yr = 2000 + int(m.group(2))
    end_dates = {1: "03-31", 2: "06-30", 3: "09-30", 4: "12-31"}
    return f"{yr}-{end_dates[q]}"


def _filename_to_quarter(filename: str) -> str:
    """Extrai trimestre do nome do arquivo.
    Ex: 'earnings-supplement-4Q25.xlsx' -> '4Q25'
    """
    m = re.search(r"(\d)Q(\d{2})", filename, re.IGNORECASE)
    if m:
        return f"{m.group(1)}Q{m.group(2)}"
    return ""


def _get_cell(ws, row: int, col: int):
    """Retorna valor numérico de uma célula."""
    return _parse_number(ws.cell(row=row, column=col).value)


def _find_row(ws, label: str, col: int = 2, start: int = 1, end: int | None = None,
              exact: bool = False, col_alt: int | None = 3) -> int | None:
    """Encontra a linha que contém label na coluna col (ou col_alt).
    Se exact=True, compara por igualdade (stripped, sem footnote markers).
    Se exact=False, usa 'in' match.
    """
    max_row = end or ws.max_row
    for r in range(start, max_row + 1):
        for c in [col, col_alt] if col_alt else [col]:
            raw = ws.cell(row=r, column=c).value
            if raw is None:
                continue
            text = str(raw).strip()
            # Remove footnote markers for comparison
            clean = re.sub(r'\s*\([a-z]\)\s*', '', text).strip()
            clean = re.sub(r'\s*\(".*?"\)\s*', '', clean).strip()
            if exact:
                if clean == label:
                    return r
            else:
                if label in clean:
                    return r
    return None


# ---------------------------------------------------------------------------
# Extração por seção
# ---------------------------------------------------------------------------

DATA_COL = 6  # Column with the current quarter's data (always col F)


def _parse_avg_balances(wb) -> dict:
    """Parse Page 6: CONDENSED AVERAGE BALANCE SHEETS AND ANNUALIZED YIELDS."""
    ws = wb["Page 6"]
    result = {"avg_balances": {}, "yields_rates": {}}
    ab = result["avg_balances"]
    yr = result["yields_rates"]

    # --- AVERAGE BALANCES (rows ~9-20) ---
    # Loans (row 14 in 4Q25)
    r = _find_row(ws, "Loans", start=8, end=20, exact=True)
    if r:
        ab["avg_loans"] = _get_cell(ws, r, DATA_COL)

    # Total interest-earning assets
    r = _find_row(ws, "Total interest-earning assets", start=8, end=25)
    if r:
        ab["avg_earning_assets"] = _get_cell(ws, r, DATA_COL)

    # TOTAL ASSETS
    r = _find_row(ws, "TOTAL ASSETS", start=8, end=25)
    if r:
        ab["avg_total_assets"] = _get_cell(ws, r, DATA_COL)

    # Interest-bearing deposits (liabilities side, row ~22)
    r = _find_row(ws, "Interest-bearing deposits", start=20, end=35, exact=True)
    if r:
        ab["avg_ib_deposits"] = _get_cell(ws, r, DATA_COL)

    # Noninterest-bearing deposits (row ~30)
    r = _find_row(ws, "Noninterest-bearing deposits", start=25, end=40, exact=True)
    if r:
        ab["avg_nib_deposits"] = _get_cell(ws, r, DATA_COL)

    # Total deposits = IB + NIB
    ib = ab.get("avg_ib_deposits")
    nib = ab.get("avg_nib_deposits")
    ab["avg_total_deposits"] = (ib + nib) if (ib is not None and nib is not None) else None

    # Total interest-bearing liabilities
    r = _find_row(ws, "Total interest-bearing liabilities", start=25, end=40)
    if r:
        ab["avg_total_ib_liabilities"] = _get_cell(ws, r, DATA_COL)

    # --- AVERAGE RATES (rows ~42-62) ---
    # Find the AVERAGE RATES section start
    rates_start = _find_row(ws, "AVERAGE RATES", start=35, end=50)
    if not rates_start:
        rates_start = 40

    # Total interest-earning assets rate
    r = _find_row(ws, "Total interest-earning assets", start=rates_start, end=rates_start + 20)
    if r:
        v = _get_cell(ws, r, DATA_COL)
        yr["avg_earning_assets_yield"] = _pct_to_decimal(v)

    # Loans rate
    r = _find_row(ws, "Loans", start=rates_start, end=rates_start + 15, exact=True)
    if r:
        v = _get_cell(ws, r, DATA_COL)
        yr["avg_loans_yield"] = _pct_to_decimal(v)

    # Interest-bearing deposits rate
    r = _find_row(ws, "Interest-bearing deposits", start=rates_start + 8, end=rates_start + 25, exact=True)
    if r:
        v = _get_cell(ws, r, DATA_COL)
        yr["avg_ib_deposits_rate"] = _pct_to_decimal(v)

    # Total interest-bearing liabilities rate
    r = _find_row(ws, "Total interest-bearing liabilities", start=rates_start + 8, end=rates_start + 25)
    if r:
        v = _get_cell(ws, r, DATA_COL)
        yr["avg_total_ib_liabilities_rate"] = _pct_to_decimal(v)

    # INTEREST RATE SPREAD
    r = _find_row(ws, "INTEREST RATE SPREAD", start=rates_start + 15, end=rates_start + 30)
    if r:
        v = _get_cell(ws, r, DATA_COL)
        yr["interest_spread"] = _pct_to_decimal(v)

    # NET YIELD ON INTEREST-EARNING ASSETS = NIM
    r = _find_row(ws, "NET YIELD ON INTEREST-EARNING ASSETS", start=rates_start + 15, end=rates_start + 30)
    if r:
        v = _get_cell(ws, r, DATA_COL)
        yr["nim"] = _pct_to_decimal(v)

    return result


def _parse_capital(wb) -> dict:
    """Parse Pages 2, 9: Capital ratios."""
    result = {}

    # --- Page 9: CAPITAL AND OTHER SELECTED BALANCE SHEET ITEMS ---
    ws9 = wb["Page 9"]

    # Standardized section (rows ~11-18)
    std_row = _find_row(ws9, "Standardized", start=9, end=15, exact=True)
    if not std_row:
        std_row = 11  # fallback

    # CET1 capital
    r = _find_row(ws9, "CET1 capital", start=std_row, end=std_row + 10, exact=True)
    if r:
        result["cet1_capital"] = _get_cell(ws9, r, DATA_COL)

    # RWA
    r = _find_row(ws9, "Risk-weighted assets", start=std_row, end=std_row + 10, exact=True)
    if r:
        result["rwa_standardized"] = _get_cell(ws9, r, DATA_COL)

    # CET1 ratio
    r = _find_row(ws9, "CET1 capital ratio", start=std_row, end=std_row + 10, exact=True)
    if r:
        v = _get_cell(ws9, r, DATA_COL)
        if v is not None:
            # Already as decimal (0.145) or as percent (14.5)?
            result["cet1_ratio"] = round(v, 6) if v < 1 else round(v / 100, 6)

    # Total capital ratio
    r = _find_row(ws9, "Total capital ratio", start=std_row, end=std_row + 10, exact=True)
    if r:
        v = _get_cell(ws9, r, DATA_COL)
        if v is not None:
            result["total_capital_ratio"] = round(v / 100, 6) if v > 1 else round(v, 6)

    # Tier 1 leverage ratio
    r = _find_row(ws9, "Tier 1 leverage ratio", start=std_row + 5, end=std_row + 30)
    if r:
        v = _get_cell(ws9, r, DATA_COL)
        if v is not None:
            result["leverage_ratio"] = round(v / 100, 6) if v > 1 else round(v, 6)

    # SLR
    r = _find_row(ws9, "SLR", start=std_row + 5, end=std_row + 30, exact=True)
    if r:
        v = _get_cell(ws9, r, DATA_COL)
        if v is not None:
            result["slr"] = round(v / 100, 6) if v > 1 else round(v, 6)

    # --- Page 2: CET1 ratio and SLR from CAPITAL RATIOS section ---
    # (already got from Page 9, use as fallback/validation)

    # LCR / NSFR - preenchido depois via _enrich_lcr_from_presentations
    result["lcr"] = None
    result["nsfr"] = None

    return result


def _extract_lcr_from_presentation(html_path: str) -> float | None:
    """Extrai Firm LCR de um earnings presentation HTML (8-K slide deck)."""
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            content = f.read()
        text = re.sub(r"<[^>]+>", " ", content)
        text = re.sub(r"&[^;]+;", " ", text)
        text = re.sub(r"\s+", " ", text)
        match = re.search(r"Firm\s+LCR\s+(\d+)\s*%", text)
        if match:
            return int(match.group(1)) / 100  # ex: 111% -> 1.11
    except Exception:
        pass
    return None


def _enrich_lcr_from_presentations(resultados: list[dict], pasta_docs: str):
    """Enriquece supplement com LCR extraído dos earnings presentations HTML."""
    # Mapear trimestre -> HTML path
    import glob as _glob
    htms = _glob.glob(os.path.join(pasta_docs, "*earningsxpresentat*.htm"))
    # Extrair trimestre do nome: a4q25 -> 4Q25, a1q25 -> 1Q25
    for htm_path in htms:
        fname = os.path.basename(htm_path)
        m = re.search(r"a(\d)q(\d{2})", fname)
        if not m:
            continue
        q = m.group(1)
        yr = m.group(2)
        trimestre = f"{q}Q{yr}"
        periodo = _quarter_to_date(trimestre)

        lcr_val = _extract_lcr_from_presentation(htm_path)
        if lcr_val is None:
            continue

        # Encontrar a entry correspondente e atualizar
        for entry in resultados:
            if entry.get("periodo") == periodo:
                cap = entry.get("capital", {})
                cap["lcr"] = lcr_val
                entry["capital"] = cap
                break


def _parse_credit_quality(wb) -> dict:
    """Parse Pages 25, 26, 27: Credit Quality."""
    result = {}

    # --- Page 25: NONPERFORMING ASSETS ---
    ws25 = wb["Page 25"]
    r = _find_row(ws25, "Total nonperforming assets", start=8, end=30)
    if r:
        result["npa"] = _get_cell(ws25, r, DATA_COL)

    # --- Page 26: NET CHARGE-OFFS and ALLOWANCE ---
    ws26 = wb["Page 26"]

    # Net charge-offs (total)
    r = _find_row(ws26, "Net charge-offs", start=10, end=20, exact=True)
    if r:
        result["nco_total"] = _get_cell(ws26, r, DATA_COL)

    # Allowance for loan losses - ending balance
    r = _find_row(ws26, "Ending balance", start=10, end=20)
    if r:
        result["acl_loans"] = _get_cell(ws26, r, DATA_COL)

    # Total allowance for credit losses
    r = _find_row(ws26, "Total allowance for credit losses", start=20, end=35)
    if r:
        result["acl_total"] = _get_cell(ws26, r, DATA_COL)

    # --- Page 27: CREDIT RATIOS ---
    ws27 = wb["Page 27"]

    # Total allowance to total retained loans
    r = _find_row(ws27, "Total allowance to total retained loans", start=25, end=40)
    if r:
        v = _get_cell(ws27, r, DATA_COL)
        if v is not None:
            result["acl_pct_loans"] = round(v / 100, 6)

    return result


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------

def extrair_supplement_jpm(pasta_docs: str, pasta_destino: str) -> list[dict]:
    """Extrai dados de todos os Earnings Supplement XLSX do JPMorgan Chase.

    Args:
        pasta_docs: Caminho para pasta com os XLSX (earnings-supplement-*.xlsx)
        pasta_destino: Caminho para pasta de destino do JSON

    Returns:
        Lista de dicionários com dados extraídos por trimestre.
    """
    xlsxs = sorted(glob.glob(os.path.join(pasta_docs, "earnings-supplement-*.xlsx")))
    if not xlsxs:
        print(f"[WARN] Nenhum XLSX encontrado em {pasta_docs}")
        return []

    resultados = []

    for xlsx_path in xlsxs:
        fname = os.path.basename(xlsx_path)
        trimestre = _filename_to_quarter(fname)
        periodo = _quarter_to_date(trimestre)

        print(f"  Processando {fname} ({trimestre})...")

        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception as e:
            print(f"    [ERRO] Não conseguiu abrir: {e}")
            continue

        entry = {
            "periodo": periodo,
            "trimestre": trimestre,
            "fonte": fname,
        }

        # Average Balances & Yields
        try:
            avg_data = _parse_avg_balances(wb)
            entry["avg_balances"] = avg_data.get("avg_balances", {})
            entry["yields_rates"] = avg_data.get("yields_rates", {})
        except Exception as e:
            print(f"    [WARN] Erro avg_balances: {e}")
            entry["avg_balances"] = {}
            entry["yields_rates"] = {}

        # Capital
        try:
            entry["capital"] = _parse_capital(wb)
        except Exception as e:
            print(f"    [WARN] Erro capital: {e}")
            entry["capital"] = {}

        # Credit Quality
        try:
            entry["credit_quality"] = _parse_credit_quality(wb)
        except Exception as e:
            print(f"    [WARN] Erro credit_quality: {e}")
            entry["credit_quality"] = {}

        wb.close()
        resultados.append(entry)

    # Enriquecer com LCR dos earnings presentations (HTMLs)
    _enrich_lcr_from_presentations(resultados, pasta_docs)

    # Preencher LCR faltante via 10-Q/10-K filings da EDGAR
    missing_lcr = [r["periodo"] for r in resultados
                   if r.get("capital", {}).get("lcr") is None]
    if missing_lcr:
        print("  Extraindo LCR dos 10-Q/10-K filings (periodos faltantes)...")
        from .extrator_supplement_bac import _extract_lcr_from_10q
        lcr_map = _extract_lcr_from_10q("JPM", missing_lcr)
        for entry in resultados:
            per = entry.get("periodo")
            if per in lcr_map:
                cap = entry.get("capital", {})
                cap["lcr"] = lcr_map[per]
                entry["capital"] = cap

    # Sort by period
    resultados.sort(key=lambda x: x["periodo"])

    # Save
    os.makedirs(pasta_destino, exist_ok=True)
    out_path = os.path.join(pasta_destino, "supplement_data.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(resultados, f, indent=2, ensure_ascii=False)

    print(f"\n  Salvo {len(resultados)} trimestres em {out_path}")
    return resultados


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    pasta_docs = sys.argv[1] if len(sys.argv) > 1 else "G:/Meu Drive/Análise de Crédito Financeiras/JPM/Documentos"
    pasta_destino = sys.argv[2] if len(sys.argv) > 2 else "G:/Meu Drive/Análise de Crédito Financeiras/JPM/Dados_EDGAR"

    dados = extrair_supplement_jpm(pasta_docs, pasta_destino)

    if dados:
        print("\n=== Últimos 2 trimestres ===")
        for entry in dados[-2:]:
            print(json.dumps(entry, indent=2, ensure_ascii=False))
