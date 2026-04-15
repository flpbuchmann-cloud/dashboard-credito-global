"""
Extrator de dados financeiros da ArcelorMittal a partir do modelo XLSM
de equity research (424090.MT.AS.XLSM).

Vantagens vs. earnings releases em PDF:
  - Cobertura quarterly de 1Q08 a 4Q25 (vs. PDF que so vai ate 2Q25)
  - Inclui linha "Interest paid" no CFS
  - Tem cronograma de divida detalhado por instrumento
  - Valores ja em USD milhoes

Saidas em pasta_destino:
  - contas_chave.json     (formato flat list, ITR_*)
  - cronogramas.json      (lista, do Debt Profile)
  - supplement_data.json  (mantido para compatibilidade)
  - ratings.json
  - ri_website.json
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl


# ---------------------------------------------------------------------------
# Mapeamento (linha XLSM -> chave contas_chave)
# Cada sheet eh acessada por nome de aba; linhas sao 1-indexed.
# ---------------------------------------------------------------------------
IS_ROWS: dict[str, int] = {
    "receita_liquida": 5,           # Sales
    "custo": 7,                     # Cost of Sales (excl. depreciation & impairment)
    "ebitda": 13,                   # Calculated Reported EBITDA
    "depreciacao_amortizacao": 21,  # Depreciation
    "ebit": 35,                     # Reported Operating Income (Loss)
    "despesas_juros": 37,           # Net interest expense
    "fx_other_fin": 38,             # Foreign exchange and other net financing
    "lucro_antes_ir": 42,           # Income before taxes
    "ir_csll": 49,                  # Income tax benefit (expense)
    "non_controlling": 53,
    "lucro_liquido": 61,            # Net income (loss) attributable to owners
}

BS_ROWS: dict[str, int] = {
    "intangivel": 5,                # Goodwill and intangible assets
    "imobilizado": 9,               # PP&E
    "investimentos_titulos": 11,    # Other investments
    "ativo_nao_circulante": 22,
    "caixa": 24,                    # Cash and cash equivalents
    "aplicacoes_financeiras_cp": 26,  # Short term investment
    "contas_a_receber": 27,
    "estoques_cp": 28,              # Inventories
    "ativo_circulante": 37,
    "ativo_total": 38,
    "patrimonio_liquido": 50,
    "emprestimos_lp": 52,           # Long-term debt
    "passivo_nao_circulante": 57,
    "emprestimos_cp": 59,           # Short-term debt
    "fornecedores": 60,             # Trade accounts payable
    "passivo_circulante": 72,
    "passivo_total": 74,
}

CFS_ROWS: dict[str, int] = {
    "fco": 48,                          # Net cash from operating activities
    "capex": 50,                        # Purchase of PPE and intangibles
    "fci": 57,                          # Net cash used in investing activities
    "captacao_divida_cp": 64,           # Proceeds from long-term debt
    "amortizacao_divida_cp": 65,        # Payments of short-term debt
    "amortizacao_divida_lp": 66,        # Payments of long-term debt
    "buyback": 69,
    "dividendos_pagos": 71,
    "dividendos_minorias": 72,
    "fcfin": 76,                        # Net cash from financing activities
    "juros_pagos": 31,                  # Interest paid
    "ir_pago": 18,                      # Taxes paid
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _is_quarterly_date(d) -> bool:
    if not isinstance(d, datetime):
        return False
    if d.month not in (3, 6, 9, 12):
        return False
    if d.day < 28 and d.day != 27:  # 27 = MT pre-2006 fiscal week-end
        return False
    return True


def _normalize_period(d: datetime) -> str:
    """Normaliza para fim-de-trimestre canonico (28→31 etc)."""
    last = {3: 31, 6: 30, 9: 30, 12: 31}[d.month]
    return f"{d.year:04d}-{d.month:02d}-{last:02d}"


def _coerce(v) -> Optional[float]:
    if v is None or v == "" or v is True or v is False:
        return None
    try:
        f = float(v)
        if f != f:  # NaN
            return None
        return f
    except (TypeError, ValueError):
        return None


def _identify_quarterly_columns(ws, ano_inicio: int = 2008) -> list[tuple[int, str]]:
    """Devolve [(col_index, periodo_iso), ...] para colunas Hist quarterly.
    Filtra a regiao annual (fy_label numerico) e mantem apenas colunas
    com label 'Q1-08', 'Q2-08' etc."""
    out: list[tuple[int, str]] = []
    seen: set[str] = set()
    for c in range(5, ws.max_column + 1):
        d = ws.cell(1, c).value
        fy_label = str(ws.cell(2, c).value or "")
        hist = str(ws.cell(3, c).value or "")
        if not _is_quarterly_date(d):
            continue
        if "Hist" not in hist:
            continue
        if not fy_label.startswith("Q"):
            continue  # pula colunas anuais (fy_label numerico)
        if d.year < ano_inicio:
            continue
        per = _normalize_period(d)
        if per in seen:
            continue
        seen.add(per)
        out.append((c, per))
    return out


# ---------------------------------------------------------------------------
# Parser principal
# ---------------------------------------------------------------------------
def extrair_xlsm_mt(xlsm_path: str, pasta_destino: str,
                    ano_inicio: int = 2018) -> list[dict]:
    wb = openpyxl.load_workbook(xlsm_path, data_only=True)
    ws_is = wb["IFRS_IS"]
    ws_bs = wb["IFRS_BS"]
    ws_cf = wb["IFRS_CFS"]

    cols_is = _identify_quarterly_columns(ws_is, ano_inicio)
    cols_bs = _identify_quarterly_columns(ws_bs, ano_inicio)
    cols_cf = _identify_quarterly_columns(ws_cf, ano_inicio)

    print(f"  IS quarterly cols: {len(cols_is)}  BS: {len(cols_bs)}  CFS: {len(cols_cf)}")

    contas: list[dict] = []
    periodos = sorted({p for _, p in cols_is + cols_bs + cols_cf})

    # Indexar por periodo
    is_by_per = {p: c for c, p in cols_is}
    bs_by_per = {p: c for c, p in cols_bs}
    cf_by_per = {p: c for c, p in cols_cf}

    def m(v):
        return v * 1e6 if v is not None else None

    for periodo in periodos:
        ano = int(periodo[:4])

        # ---- DRE ----
        if periodo in is_by_per:
            c = is_by_per[periodo]
            dre = {k: _coerce(ws_is.cell(r, c).value) for k, r in IS_ROWS.items()}
            # custo: o XLSM ja vem negativo (Cost of Sales)
            cost = dre.pop("custo")
            # net interest expense ja vem negativo (despesa)
            ni = dre.pop("despesas_juros")
            fx = dre.pop("fx_other_fin")
            non_ctrl = dre.pop("non_controlling")
            contas.append({"periodo": periodo, "tipo": "ITR_dre", "ano": ano, "contas": {
                "receita_liquida": m(dre.get("receita_liquida")),
                "custo": m(cost) if cost is not None else 0.0,
                "resultado_bruto": m((dre.get("receita_liquida") or 0) + (cost or 0))
                                   if dre.get("receita_liquida") is not None else None,
                "ebit": m(dre.get("ebit")),
                "ebitda": m(dre.get("ebitda")),
                "depreciacao_amortizacao": (
                    m(abs(dre.get("depreciacao_amortizacao")))
                    if dre.get("depreciacao_amortizacao") is not None else None
                ),
                "lucro_liquido": m(dre.get("lucro_liquido")),
                "lucro_antes_ir": m(dre.get("lucro_antes_ir")),
                "ir_csll": m(dre.get("ir_csll")),
                "despesas_financeiras": m(ni) if ni is not None else 0.0,
                "receitas_financeiras": 0.0,
                "resultado_financeiro": m((ni or 0) + (fx or 0)) if (ni is not None or fx is not None) else None,
            }})

        # ---- BPA ----
        if periodo in bs_by_per:
            c = bs_by_per[periodo]
            bpa = {k: _coerce(ws_bs.cell(r, c).value) for k, r in BS_ROWS.items()
                   if r <= 38 or k in ("emprestimos_lp", "passivo_nao_circulante")}
            contas.append({"periodo": periodo, "tipo": "ITR_bpa", "ano": ano, "contas": {
                "ativo_total": m(bpa.get("ativo_total")),
                "ativo_circulante": m(bpa.get("ativo_circulante")),
                "ativo_nao_circulante": m(bpa.get("ativo_nao_circulante")),
                "caixa": m(bpa.get("caixa")),
                "aplicacoes_financeiras_cp": m(bpa.get("aplicacoes_financeiras_cp")),
                "contas_a_receber": m(bpa.get("contas_a_receber")),
                "estoques_cp": m(bpa.get("estoques_cp")),
                "imobilizado": m(bpa.get("imobilizado")),
                "intangivel": m(bpa.get("intangivel")),
                "investimentos_titulos": m(bpa.get("investimentos_titulos")),
            }})

            # ---- BPP (mesma sheet, linhas diferentes) ----
            bpp_keys = ["patrimonio_liquido", "emprestimos_lp", "passivo_nao_circulante",
                        "emprestimos_cp", "fornecedores", "passivo_circulante", "passivo_total"]
            bpp = {k: _coerce(ws_bs.cell(BS_ROWS[k], c).value) for k in bpp_keys}
            contas.append({"periodo": periodo, "tipo": "ITR_bpp", "ano": ano, "contas": {
                "patrimonio_liquido": m(bpp["patrimonio_liquido"]),
                "emprestimos_cp": m(bpp["emprestimos_cp"]),
                "emprestimos_lp": m(bpp["emprestimos_lp"]),
                "fornecedores": m(bpp["fornecedores"]),
                "passivo_circulante": m(bpp["passivo_circulante"]),
                "passivo_nao_circulante": m(bpp["passivo_nao_circulante"]),
                "passivo_total": m(bpp["passivo_total"]),
            }})

        # ---- DFC ----
        if periodo in cf_by_per:
            c = cf_by_per[periodo]
            cf = {k: _coerce(ws_cf.cell(r, c).value) for k, r in CFS_ROWS.items()}
            # capex vem negativo no XLSM
            capex = cf.get("capex")
            # captacao = soma dos proceeds; amortizacao = soma dos payments
            cap_lp = cf.get("captacao_divida_cp") or 0  # nome herdado, eh LP
            amort_cp = cf.get("amortizacao_divida_cp") or 0
            amort_lp = cf.get("amortizacao_divida_lp") or 0
            captacao = cap_lp
            amortizacao = amort_cp + amort_lp
            contas.append({"periodo": periodo, "tipo": "ITR_dfc", "ano": ano, "contas": {
                "fco": m(cf.get("fco")),
                "capex": m(capex),
                "fci": m(cf.get("fci")),
                "fcf": m((cf.get("fco") or 0) + (capex or 0)) if cf.get("fco") is not None and capex is not None else None,
                "depreciacao_amortizacao": m(abs(_coerce(ws_cf.cell(12, c).value)))
                                            if _coerce(ws_cf.cell(12, c).value) is not None else None,
                "captacao_divida": m(captacao),
                "amortizacao_divida": m(amortizacao),
                "dividendos_pagos": m(cf.get("dividendos_pagos")),
                "buyback": m(cf.get("buyback")),
                "juros_pagos": m(cf.get("juros_pagos")),
                "ir_pago": m(cf.get("ir_pago")),
            }})

    # ----- Cronograma de divida do Debt Profile -----
    cronograma = _extrair_cronograma(wb)

    # ----- Salvar -----
    destino = Path(pasta_destino)
    destino.mkdir(parents=True, exist_ok=True)

    (destino / "contas_chave.json").write_text(
        json.dumps(contas, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (destino / "cronogramas.json").write_text(
        json.dumps(cronograma, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    # ratings + ri preservados (ja existentes do parser PDF)
    if not (destino / "ratings.json").exists():
        (destino / "ratings.json").write_text(json.dumps({
            "empresa": "ArcelorMittal", "ticker": "MT",
            "ratings_atuais": {
                "Moodys": {"rating": "Baa3", "outlook": "Stable"},
                "SP": {"rating": "BBB-", "outlook": "Stable"},
                "Fitch": {"rating": "BBB-", "outlook": "Stable"},
            },
            "fonte": "Public ratings agencies",
        }, indent=2, ensure_ascii=False), encoding="utf-8")
    if not (destino / "ri_website.json").exists():
        (destino / "ri_website.json").write_text(json.dumps({
            "empresa": "ArcelorMittal", "ticker": "MT",
            "ri_url": "https://corporate.arcelormittal.com/investors",
            "documentos_url": "https://corporate.arcelormittal.com/investors/financial-reports",
        }, indent=2, ensure_ascii=False), encoding="utf-8")

    return contas


def _extrair_cronograma(wb) -> list[dict]:
    """Extrai cronograma de divida da aba 'Debt Profile'.
    Agrega vencimentos por ano. Usa o ULTIMO trimestre disponivel para
    obter o saldo principal de cada instrumento."""
    if "Debt Profile" not in wb.sheetnames:
        return []
    ws = wb["Debt Profile"]

    # Encontrar coluna do ultimo trimestre Hist
    last_col = None
    last_per = None
    for c in range(5, ws.max_column + 1):
        d = ws.cell(1, c).value
        label = str(ws.cell(3, c).value or "")
        if _is_quarterly_date(d) and "Hist" in label:
            if last_col is None or d > ws.cell(1, last_col).value:
                last_col = c
                last_per = _normalize_period(d)
    if last_col is None:
        return []

    # Varrer linhas: cada instrumento tem nome em uma linha, "Principal" + maturity na proxima
    vencimentos: dict[str, float] = {}  # ano -> total USD M
    for r in range(4, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a == "Principal":
            mat = ws.cell(r, 7).value
            ano_venc = _parse_maturity(mat)
            if ano_venc is None:
                continue
            principal = _coerce(ws.cell(r, last_col).value)
            if principal is None or principal == 0:
                continue
            chave = str(ano_venc)
            vencimentos[chave] = vencimentos.get(chave, 0) + abs(principal)

    if not vencimentos:
        return []

    # Saida no formato esperado pelo dashboard
    venc_dict = {k: round(v * 1e6, 2) for k, v in sorted(vencimentos.items())}
    return [{
        "data_referencia": last_per,
        "fonte": "ArcelorMittal Debt Profile (XLSM equity research model)",
        "moeda": "USD",
        "vencimentos": venc_dict,
    }]


def _parse_maturity(mat) -> Optional[int]:
    """Extrai ano de uma string como 'Jan 17, 2024', '2025', '2025-2027'."""
    if mat is None:
        return None
    if isinstance(mat, (int, float)):
        return int(mat)
    if isinstance(mat, datetime):
        return mat.year
    s = str(mat).strip()
    # range '2025-2027' -> usa o primeiro
    m = re.match(r"^(\d{4})\s*[-–]\s*\d{2,4}", s)
    if m:
        return int(m.group(1))
    m = re.search(r"\b(20\d{2})\b", s)
    if m:
        return int(m.group(1))
    return None


if __name__ == "__main__":
    xlsm = r"G:\Meu Drive\Análise de Crédito Global\MT\Documentos\424090.MT.AS.XLSM"
    destino = r"G:\Meu Drive\Análise de Crédito Global\MT\Dados_EDGAR"
    print(f"Lendo XLSM: {xlsm}")
    contas = extrair_xlsm_mt(xlsm, destino, ano_inicio=2018)
    periodos = sorted({c["periodo"] for c in contas})
    print(f"\nTotal registros: {len(contas)}  |  Periodos: {len(periodos)}")
    print(f"Range: {periodos[0]} -> {periodos[-1]}")
    print(f"Salvo em: {destino}")
